# -*- coding: utf-8 -*-
"""
效率计算服务 - 从三个数据源拉数据，统一计算效率
- 按人员汇总
- 按日期计算（今日、近7天、近30天）
- 按组别分类
- 计算达标率
- 带缓存机制
"""
import json
import time
import logging
from datetime import datetime, date, timedelta
from typing import Dict, List, Optional, Tuple

from services.data_sources import GodzillaClient, MirrorClient, WatcherClient
from config import EFFICIENCY_CONFIG, STAFF_ROSTER_PATH, CACHE_TTL_SECONDS

logger = logging.getLogger(__name__)


# ==================== 员工花名册 ====================

def load_staff_roster() -> Dict[str, Dict]:
    """
    加载员工花名册，返回 {employee_no: staff_info} 的字典
    """
    try:
        with open(STAFF_ROSTER_PATH, 'r', encoding='utf-8') as f:
            data = json.load(f)
        roster = {}
        for s in data.get('staff', []):
            emp_no = str(s.get('employee_no', ''))
            if not emp_no:
                emp_no = 'NAME_' + s.get('employee_name', 'unknown')
            roster[emp_no] = s
        return roster
    except Exception as e:
        logger.error(f'Load staff roster failed: {e}')
        return {}


def get_staff_info(employee_no: str, roster: Dict[str, Dict] = None) -> Dict:
    """获取员工信息"""
    if roster is None:
        roster = load_staff_roster()
    return roster.get(str(employee_no), {})


def get_staff_group(employee_no: str, roster: Dict[str, Dict] = None) -> str:
    """获取员工所属组别"""
    info = get_staff_info(employee_no, roster)
    return info.get('dept_group', '未分组')


def get_staff_name(employee_no: str, roster: Dict[str, Dict] = None) -> str:
    """获取员工姓名"""
    info = get_staff_info(employee_no, roster)
    return info.get('employee_name', employee_no)


def get_target_efficiency(employee_no: str, data_source: str,
                          roster: Dict[str, Dict] = None) -> float:
    """
    获取员工的目标效率（UPPH标准）
    根据员工类型（全职/兼职）和数据源不同，标准不同
    """
    info = get_staff_info(employee_no, roster)
    emp_type = info.get('employee_type', '')

    if data_source == 'godzilla' or data_source.startswith('watcher_godzilla'):
        if '全职' in emp_type:
            return EFFICIENCY_CONFIG['fulltime_standard']
        else:
            return EFFICIENCY_CONFIG['parttime_standard']
    elif data_source == 'mirror':
        if '全职' in emp_type:
            return EFFICIENCY_CONFIG['mirror_fulltime_standard']
        else:
            return EFFICIENCY_CONFIG['mirror_parttime_standard']
    else:
        return 100.0


# ==================== 缓存机制 ====================

class CacheStore:
    """简单的内存缓存"""

    def __init__(self, ttl: int = CACHE_TTL_SECONDS):
        self._cache: Dict[str, Tuple[float, any]] = {}
        self.ttl = ttl

    def get(self, key: str) -> Optional[any]:
        """获取缓存，过期返回None"""
        if key not in self._cache:
            return None
        timestamp, value = self._cache[key]
        if time.time() - timestamp > self.ttl:
            del self._cache[key]
            return None
        return value

    def set(self, key: str, value: any):
        """设置缓存"""
        self._cache[key] = (time.time(), value)

    def invalidate(self, prefix: str = None):
        """清除缓存"""
        if prefix:
            keys_to_del = [k for k in self._cache if k.startswith(prefix)]
            for k in keys_to_del:
                del self._cache[k]
        else:
            self._cache.clear()


# 全局缓存实例
_cache = CacheStore()


def invalidate_cache(prefix: str = None):
    """清除缓存"""
    _cache.invalidate(prefix)


# ==================== 效率计算服务 ====================

class EfficiencyService:
    """
    统一的效率计算服务
    从哥斯拉、魔镜、Watcher三个数据源拉取数据，
    进行汇总、计算UPPH、达标率等指标
    """

    def __init__(self):
        self.godzilla_client = GodzillaClient()
        self.mirror_client = MirrorClient()
        self.watcher_client = WatcherClient()
        self.roster = load_staff_roster()

    def _get_cache_key(self, prefix: str, *args) -> str:
        """生成缓存key"""
        return f'{prefix}:' + ':'.join(str(a) for a in args)

    # ---------- 数据获取（带缓存） ----------

    def get_godzilla_records(self, target_date: str) -> List[Dict]:
        """获取哥斯拉当日记录（带缓存）"""
        cache_key = self._get_cache_key('godzilla_records', target_date)
        cached = _cache.get(cache_key)
        if cached is not None:
            return cached

        records = self.godzilla_client.fetch_by_date(target_date)
        _cache.set(cache_key, records)
        return records

    def get_mirror_records(self, target_date: str) -> List[Dict]:
        """获取魔镜当日记录（带缓存）"""
        cache_key = self._get_cache_key('mirror_records', target_date)
        cached = _cache.get(cache_key)
        if cached is not None:
            return cached

        records = self.mirror_client.fetch_by_date(target_date)
        _cache.set(cache_key, records)
        return records

    # ---------- 按员工汇总 ----------

    def _aggregate_by_staff(self, records: List[Dict], data_source: str) -> Dict[str, Dict]:
        """
        将明细记录按员工汇总
        返回: {employee_no: {employee_no, employee_name, total_count,
                             total_duration_sec, work_hours, efficiency, ...}}
        """
        staff_map: Dict[str, Dict] = {}

        for rec in records:
            emp_no = str(rec.get('employee_no', ''))
            if not emp_no or emp_no == 'None':
                continue

            if emp_no not in staff_map:
                name = rec.get('employee_name') or get_staff_name(emp_no, self.roster)
                group = get_staff_group(emp_no, self.roster)
                staff_map[emp_no] = {
                    'employee_no': emp_no,
                    'employee_name': name,
                    'group_name': group,
                    'total_count': 0,
                    'total_duration_sec': 0,
                    'first_time': None,
                    'last_time': None,
                    'data_source': data_source,
                    'position_tag': rec.get('position_tag', ''),
                    'machine_ids': set(),
                }

            s = staff_map[emp_no]

            # 累加数量 - 每条记录代表1个产品/1次质检，UPPH按产品数计算
            # 同时记录photo_count用于参考展示
            s['total_count'] += 1  # 产品数/质检次数次数
            # 收集设备编号
            m_id = rec.get('machine_id', '') or ''
            if m_id:
                s['machine_ids'].add(str(m_id))
            if 'photo_count' in rec and rec['photo_count']:
                s['total_photos'] = s.get('total_photos', 0) + rec['photo_count']

            # 累加工时
            dur = rec.get('total_duration_sec', 0) or rec.get('duration_sec', 0) or 0
            s['total_duration_sec'] += int(dur)

            # 记录首次和末次时间（用于计算在岗时长）
            start_time = rec.get('start_time') or rec.get('date', '')
            if start_time:
                try:
                    dt = datetime.strptime(start_time[:19], '%Y-%m-%d %H:%M:%S')
                    if s['first_time'] is None or dt < s['first_time']:
                        s['first_time'] = dt
                    if s['last_time'] is None or dt > s['last_time']:
                        s['last_time'] = dt
                except (ValueError, IndexError):
                    pass

        # 计算工作时长和效率
        for emp_no, s in staff_map.items():
            # 方法1: 用总耗时长 / 3600
            duration_hours = s['total_duration_sec'] / 3600.0 if s['total_duration_sec'] > 0 else 0

            # 工作时长 = 首末次时间差 - 午休（参考efficiency_sync.ps1的Calc-EfficiencyDuration）
            # 设备累计时长仅作参考，不作为UPPH计算依据
            work_hours = 0
            if s['first_time'] and s['last_time']:
                span = s['last_time'] - s['first_time']
                span_hours = span.total_seconds() / 3600.0
                # 减去午休
                lunch_hour = EFFICIENCY_CONFIG['lunch_break_hour']
                lunch_dur = EFFICIENCY_CONFIG['lunch_break_duration_hours']
                if s['last_time'].hour >= lunch_hour and s['first_time'].hour < lunch_hour + 1:
                    span_hours = max(0, span_hours - lunch_dur)
                work_hours = span_hours
                s['span_hours'] = round(span_hours, 2)
            else:
                s['span_hours'] = 0

            s['work_hours'] = round(work_hours, 2)
            s['duration_hours'] = round(duration_hours, 2)

            # 计算UPPH (Units Per Person Per Hour)
            if work_hours > 0:
                s['efficiency'] = round(s['total_count'] / work_hours, 1)
            else:
                s['efficiency'] = 0

            # 计算达标率
            target = get_target_efficiency(emp_no, data_source, self.roster)
            s['target_efficiency'] = target
            if target > 0 and s['efficiency'] > 0:
                s['achievement_rate'] = round(s['efficiency'] / target * 100, 1)
            else:
                s['achievement_rate'] = 0

            # 设备编号：去重后取前5个，用逗号分隔
            if s.get('machine_ids'):
                sorted_machines = sorted(s['machine_ids'])
                s['machine_id'] = ', '.join(sorted_machines[:5])
                if len(sorted_machines) > 5:
                    s['machine_id'] += f' 等{len(sorted_machines)}个'
            else:
                s['machine_id'] = ''
            # 删除set避免JSON序列化问题
            if 'machine_ids' in s:
                del s['machine_ids']

            # 转换datetime为字符串（JSON序列化需要）
            if s.get('first_time'):
                s['first_time'] = s['first_time'].strftime('%Y-%m-%d %H:%M:%S')
            if s.get('last_time'):
                s['last_time'] = s['last_time'].strftime('%Y-%m-%d %H:%M:%S')

        return staff_map

    # ---------- 按组别汇总 ----------

    def _aggregate_by_group(self, staff_map: Dict[str, Dict]) -> Dict[str, Dict]:
        """
        将员工数据按组别汇总
        """
        group_map: Dict[str, Dict] = {}

        for emp_no, staff in staff_map.items():
            group = staff['group_name']
            if group not in group_map:
                group_map[group] = {
                    'group_name': group,
                    'total_count': 0,
                    'total_work_hours': 0,
                    'staff_count': 0,
                    'achievement_rates': [],
                    'efficiencies': [],
                    'data_source': staff['data_source'],
                }

            g = group_map[group]
            g['total_count'] += staff['total_count']
            g['total_work_hours'] += staff['work_hours']
            g['staff_count'] += 1
            if staff['efficiency'] > 0:
                g['efficiencies'].append(staff['efficiency'])
            if staff['achievement_rate'] > 0:
                g['achievement_rates'].append(staff['achievement_rate'])

        # 计算组级别指标
        for group, g in group_map.items():
            if g['total_work_hours'] > 0:
                g['avg_efficiency'] = round(g['total_count'] / g['total_work_hours'], 1)
            else:
                g['avg_efficiency'] = 0

            if g['achievement_rates']:
                g['achievement_rate'] = round(
                    sum(g['achievement_rates']) / len(g['achievement_rates']), 1
                )
            else:
                g['achievement_rate'] = 0

            # 达标人数
            g['achieved_count'] = sum(1 for r in g['achievement_rates'] if r >= 100)
            g['total_work_hours'] = round(g['total_work_hours'], 1)

        return group_map

    # ---------- 对外API: 单天数据 ----------

    def get_photo_efficiency(self, target_date: str = None) -> Dict:
        """
        获取拍照效率（哥斯拉数据）
        """
        if target_date is None:
            target_date = date.today().strftime('%Y-%m-%d')

        cache_key = self._get_cache_key('photo_efficiency', target_date)
        cached = _cache.get(cache_key)
        if cached is not None:
            return cached

        records = self.get_godzilla_records(target_date)
        staff_map = self._aggregate_by_staff(records, 'godzilla')
        group_map = self._aggregate_by_group(staff_map)

        # 按达标率排序
        staff_list = sorted(
            staff_map.values(),
            key=lambda x: x['achievement_rate'],
            reverse=True
        )

        # 标记前三名
        for i, s in enumerate(staff_list):
            s['rank'] = i + 1
            s['is_top'] = (i == 0 and s['achievement_rate'] > 0)

        # 整体统计
        total_count = sum(s['total_count'] for s in staff_list)
        total_hours = sum(s['work_hours'] for s in staff_list)
        avg_eff = round(total_count / total_hours, 1) if total_hours > 0 else 0
        rates = [s['achievement_rate'] for s in staff_list if s['achievement_rate'] > 0]
        avg_rate = round(sum(rates) / len(rates), 1) if rates else 0

        result = {
            'date': target_date,
            'data_source': 'godzilla',
            'total_count': total_count,
            'total_work_hours': round(total_hours, 1),
            'avg_efficiency': avg_eff,
            'achievement_rate': avg_rate,
            'staff_count': len(staff_list),
            'staff_ranking': staff_list,
            'groups': list(group_map.values()),
            'group_count': len(group_map),
            'session_expired': self.godzilla_client.session_expired,
        }

        _cache.set(cache_key, result)
        return result

    def get_automation_efficiency(self, target_date: str = None) -> Dict:
        """
        获取自动化效率（哥斯拉+魔镜双数据源）
        - 哥斯拉系统的数据 → 哥斯拉组
        - 魔镜系统的数据 → 魔镜组
        - 人员按花名册position_tag筛选（只保留哥斯拉/魔镜标签的人）
        - 同一人在两边都有数据时，各算各的，分别在两个组里展示
        """
        if target_date is None:
            target_date = date.today().strftime('%Y-%m-%d')

        cache_key = self._get_cache_key('automation_efficiency', target_date)
        cached = _cache.get(cache_key)
        if cached is not None:
            return cached

        # 同时拉取哥斯拉和魔镜数据
        godzilla_records = self.get_godzilla_records(target_date)
        mirror_records = self.get_mirror_records(target_date)

        # 分别聚合（_aggregate_by_staff返回{emp_no: staff_data}）
        godzilla_staff = self._aggregate_by_staff(godzilla_records, 'godzilla')
        mirror_staff = self._aggregate_by_staff(mirror_records, 'mirror')

        # 最终人员列表的key = emp_no + "_" + data_source，避免同一人两边数据互相覆盖
        auto_staff = {}

        # === 处理哥斯拉数据 ===
        for emp_no, s in godzilla_staff.items():
            info = get_staff_info(emp_no, self.roster)
            pos_tag = info.get('position_tag', '')
            # 只保留position_tag为哥斯拉/魔镜的人员
            if pos_tag not in ('哥斯拉', '魔镜'):
                continue
            # 补全信息
            s['name'] = info.get('name', s.get('name', emp_no))
            s['employee_type'] = info.get('employee_type', '')
            s['position_tag'] = pos_tag
            s['group_name'] = '哥斯拉'
            s['data_source'] = 'godzilla'
            # 计算达标率（哥斯拉标准：全职58、兼职52）
            emp_type = s.get('employee_type', '')
            if '全职' in emp_type:
                target = EFFICIENCY_CONFIG['fulltime_standard']
            else:
                target = EFFICIENCY_CONFIG['parttime_standard']
            s['target_efficiency'] = target
            if target > 0 and s.get('efficiency', 0) > 0:
                s['achievement_rate'] = round(s['efficiency'] / target * 100, 1)
            else:
                s['achievement_rate'] = 0
            key = emp_no + '_godzilla'
            auto_staff[key] = s

        # === 处理魔镜数据 ===
        for emp_no, s in mirror_staff.items():
            info = get_staff_info(emp_no, self.roster)
            pos_tag = info.get('position_tag', '')
            # 只保留position_tag为哥斯拉/魔镜的人员
            if pos_tag not in ('哥斯拉', '魔镜'):
                continue
            # 补全信息
            s['name'] = info.get('name', s.get('name', emp_no))
            s['employee_type'] = info.get('employee_type', '')
            s['position_tag'] = pos_tag
            s['group_name'] = '魔镜'
            s['data_source'] = 'mirror'
            # 计算达标率（魔镜统一183.7）
            target = EFFICIENCY_CONFIG['mirror_fulltime_standard']
            s['target_efficiency'] = target
            if target > 0 and s.get('efficiency', 0) > 0:
                s['achievement_rate'] = round(s['efficiency'] / target * 100, 1)
            else:
                s['achievement_rate'] = 0
            key = emp_no + '_mirror'
            auto_staff[key] = s

        # === 组装结果 ===
        staff_list = sorted(
            auto_staff.values(),
            key=lambda x: x.get('achievement_rate', 0),
            reverse=True
        )
        for i, s in enumerate(staff_list):
            s['rank'] = i + 1
            s['is_top'] = (i == 0 and s.get('achievement_rate', 0) > 0)

        # 汇总
        total_count = sum(s.get('total_count', 0) for s in staff_list)
        total_hours = sum(s.get('work_hours', 0) for s in staff_list)
        avg_eff = round(total_count / total_hours, 1) if total_hours > 0 else 0
        rates = [s['achievement_rate'] for s in staff_list if s.get('achievement_rate', 0) > 0]
        avg_rate = round(sum(rates) / len(rates), 1) if rates else 0

        # 分组统计
        groups = []
        for group_name in ['哥斯拉', '魔镜']:
            group_staff = [s for s in staff_list if s.get('group_name') == group_name]
            g_count = sum(s.get('total_count', 0) for s in group_staff)
            g_hours = sum(s.get('work_hours', 0) for s in group_staff)
            g_avg_eff = round(g_count / g_hours, 1) if g_hours > 0 else 0
            g_rates = [s['achievement_rate'] for s in group_staff if s.get('achievement_rate', 0) > 0]
            g_avg_rate = round(sum(g_rates) / len(g_rates), 1) if g_rates else 0
            g_achieved = sum(1 for s in group_staff if s.get('achievement_rate', 0) >= 100)
            groups.append({
                'group_name': group_name,
                'total_count': g_count,
                'staff_count': len(group_staff),
                'achieved_count': g_achieved,
                'achievement_rate': g_avg_rate,
                'avg_efficiency': g_avg_eff,
                'total_work_hours': round(g_hours, 1),
            })

        session_expired = self.godzilla_client.session_expired or self.mirror_client.session_expired

        result = {
            'date': target_date,
            'data_source': 'godzilla+mirror',
            'total_count': total_count,
            'staff_count': len(staff_list),
            'avg_efficiency': avg_eff,
            'achievement_rate': avg_rate,
            'total_work_hours': round(total_hours, 1),
            'staff_list': staff_list,
            'staff_ranking': staff_list,
            'groups': groups,
            'session_expired': session_expired,
        }

        _cache.set(cache_key, result)
        return result


    def _recalc_efficiency(self, s, pos_tag):
        """重新计算UPPH和达标率（按position_tag对应的标准）"""
        # 工作时长已经在_aggregate_by_staff里算好了，直接用
        work_hours = s.get('work_hours', 0)
        total_count = s.get('total_count', 0)
        if work_hours > 0:
            s['efficiency'] = round(total_count / work_hours, 1)
        else:
            s['efficiency'] = 0
        # 达标标准
        emp_type = s.get('employee_type', '')
        if pos_tag == '魔镜':
            target = EFFICIENCY_CONFIG['mirror_fulltime_standard']  # 魔镜统一183.7
        elif '全职' in emp_type:
            target = EFFICIENCY_CONFIG['fulltime_standard']
        else:
            target = EFFICIENCY_CONFIG['parttime_standard']
        s['target_efficiency'] = target
        if target > 0 and s['efficiency'] > 0:
            s['achievement_rate'] = round(s['efficiency'] / target * 100, 1)
        else:
            s['achievement_rate'] = 0


    def get_app_efficiency(self, target_date=None):
        """获取APP效率数据 - 从金山文档v3 API下载xlsx解析"""
        import traceback
        from datetime import datetime, timedelta
        
        result = {
            'date': target_date or datetime.now().strftime('%Y-%m-%d'),
            'total_count': 0,
            'staff_count': 0,
            'avg_efficiency': 0,
            'achievement_rate': 0,
            'total_work_hours': 0,
            'staff_ranking': [],
            'source': 'kdocs',
            'status': 'error',
            'message': '',
        }
        
        try:
            if target_date is None:
                target_date = datetime.now().strftime('%Y-%m-%d')
            result['date'] = target_date
            
            # 初始化金山文档客户端
            if not hasattr(self, '_kdocs') or self._kdocs is None:
                from services.kdocs_client import KdocsClient
                from config import KDOCS_CONFIG
                self._kdocs = KdocsClient(
                    token=KDOCS_CONFIG.get('token', ''),
                    folder_share_url=KDOCS_CONFIG.get('folder_share_url', '')
                )
                cookie = KDOCS_CONFIG.get('cookie', '')
                if cookie:
                    if hasattr(self._kdocs, 'set_cookie'):
                        self._kdocs.set_cookie(cookie)
                    else:
                        self._kdocs.session.headers['Cookie'] = cookie
                teamid = KDOCS_CONFIG.get('teamid', '')
                if teamid and hasattr(self._kdocs, 'set_teamid'):
                    self._kdocs.set_teamid(teamid)
            
            # 列文件
            if hasattr(self._kdocs, 'list_team_files'):
                files = self._kdocs.list_team_files(limit=200)
                logger.info(f'[APP效率] 金山文档文件列表: {len(files)}个')
            else:
                files = []
                logger.warning('[APP效率] kdocs客户端没有list_team_files方法')
            
            if not files:
                result['status'] = 'no_files'
                result['message'] = '未获取到文件列表'
                return result
            
            xlsx_files = [f for f in files if (f.get('fname') or f.get('name','')).endswith(('.xlsx','.xls','.et'))]
            logger.info(f'[APP效率] xlsx文件数: {len(xlsx_files)}个')
            
            import openpyxl
            staff_list = []
            
            for _idx, f in enumerate(xlsx_files):
                fname = f.get('fname') or f.get('name', '')
                file_id = f.get('id')
                if not file_id:
                    continue
                try:
                    logger.info(f'[APP效率] 处理 {_idx+1}/{len(xlsx_files)}: {fname[:20]}')
                    import signal
                    dl_path = None
                    try:
                        dl_path = self._kdocs.download_xlsx(file_id)
                    except Exception as dl_e:
                        logger.warning(f'[APP效率] 下载失败 {fname}: {dl_e}')
                        continue
                    if not dl_path:
                        logger.warning(f'[APP效率] 下载失败: {fname[:20]}')
                        continue
                    person_data = self._parse_app_xlsx(dl_path)
                    if not person_data.get('staff_name') or person_data.get('staff_name', '').startswith('kdocs') or person_data.get('staff_name') == 'kdocs':
                        name = fname.replace('.xlsx','').replace('.xls','').replace('.et','')
                        if '——' in name:
                            name = name.split('——')[0]
                        elif '-' in name:
                            name = name.split('-')[0]
                        person_data['staff_name'] = name.strip()
                    staff_list.append(person_data)
                    try:
                        import os
                        os.remove(dl_path)
                    except:
                        pass
                except Exception as e:
                    print(f"[app_eff] 处理 {fname} 失败: {e}")
                    continue
            
            logger.info(f'[APP效率] 成功解析 {len(staff_list)} 人数据')
            today_count = sum(s.get('today_count', 0) for s in staff_list)
            month_count = sum(s.get('month_count', 0) for s in staff_list)
            
            staff_list.sort(key=lambda x: x.get('today_count', 0), reverse=True)
            staff_ranking = []
            for s in staff_list:
                staff_ranking.append({
                    'name': s.get('staff_name', ''),
                    'today': s.get('today_count', 0),
                    'month': s.get('month_count', 0),
                    'total': s.get('total_records', 0),
                    'upph': s.get('upph', 0),
                    'achievement_rate': 0,
                    'work_hours': 0,
                })
            
            result['total_count'] = today_count
            result['staff_count'] = len(staff_list)
            result['avg_efficiency'] = round(today_count / len(staff_list), 1) if staff_list else 0
            result['staff_ranking'] = staff_ranking
            result['today_count'] = today_count
            result['month_count'] = month_count
            result['status'] = 'ok'
            return result
            
        except Exception as e:
            traceback.print_exc()
            result['status'] = 'error'
            result['message'] = str(e)
            return result

    def _parse_app_xlsx(self, xlsx_path):
        """解析单个APP登记xlsx文件 - 表头定位法
        先找表头行（含"扫码时间"/"日期"文字），根据表头定位列，再读数据
        """
        import os as _os
        import openpyxl
        from datetime import datetime, date, time
        
        result = {
            'staff_name': '',
            'today_count': 0,
            'month_count': 0,
            'yesterday_count': 0,
            'total_records': 0,
            'upph': 0,
            'android_count': 0,
            'ios_count': 0,
            'total_count': 0,
            'android_first': None,
            'android_last': None,
            'ios_first': None,
            'ios_last': None,
            'overall_first': None,
            'overall_last': None,
            'android_hours': 0,
            'ios_hours': 0,
            'work_hours': 0,
            'work_type': '安卓',
        }
        
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            
            # 从文件名提取员工姓名（排除kdocs临时文件）
            fname_base = _os.path.basename(xlsx_path)
            name_extracted = False
            for sep in ['——', '--', '-', '_']:
                if sep in fname_base:
                    candidate = fname_base.split(sep)[0].strip()
                    if not candidate.startswith('kdocs_') and len(candidate) > 1 and candidate != 'kdocs':
                        result['staff_name'] = candidate
                        name_extracted = True
                        break
            
            # 从A1单元格读姓名兜底
            if not name_extracted:
                try:
                    ws_first = wb[wb.sheetnames[0]]
                    a1_val = ws_first.cell(row=1, column=1).value
                    if a1_val and isinstance(a1_val, str) and len(a1_val) < 20:
                        result['staff_name'] = a1_val.strip()
                except:
                    pass
            
            # 分类sheet
            android_sheets = []
            ios_sheets = []
            for s in wb.sheetnames:
                s_low = s.lower()
                if 'app' in s_low and ('登记' in s or '效率' in s or '统计' in s or '装跑' in s):
                    android_sheets.append(s)
                elif '苹果' in s or '测跑' in s or 'ios' in s_low:
                    ios_sheets.append(s)
            
            if not android_sheets and not ios_sheets and wb.sheetnames:
                android_sheets = [wb.sheetnames[0]]
            
            def _parse_sheet(ws):
                """解析单个sheet，返回(count, first_dt, last_dt)"""
                # 第1步：找表头行（包含"扫码时间"或"日期"字样）
                header_row = None
                datetime_col = None
                date_col = None
                
                for scan_row in range(1, min(20, ws.max_row + 1)):
                    found_any = False
                    for scan_col in range(1, min(15, ws.max_column + 1)):
                        v = ws.cell(row=scan_row, column=scan_col).value
                        if isinstance(v, str):
                            v_str = v.strip()
                            if '扫码时间' in v_str:
                                datetime_col = scan_col
                                found_any = True
                            elif v_str == '日期':
                                date_col = scan_col
                                found_any = True
                    if found_any:
                        header_row = scan_row
                        break
                
                if header_row is None:
                    header_row = 4
                    if datetime_col is None:
                        datetime_col = 3
                    if date_col is None:
                        date_col = 4
                
                start_row = header_row + 1
                count = 0
                first_dt = None
                last_dt = None
                
                for row_idx in range(start_row, ws.max_row + 1):
                    try:
                        row_dt = None
                        if datetime_col:
                            v = ws.cell(row=row_idx, column=datetime_col).value
                            if isinstance(v, datetime):
                                row_dt = v
                            elif isinstance(v, date):
                                row_dt = datetime.combine(v, time(0, 0, 0))
                            elif isinstance(v, str) and len(v) >= 10:
                                for fmt in ['%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S',
                                            '%Y-%m-%d %H:%M', '%Y/%m/%d %H:%M']:
                                    try:
                                        s = v.strip()
                                        row_dt = datetime.strptime(s[:19] if len(s) >= 19 else s, fmt)
                                        break
                                    except:
                                        pass
                        if not row_dt and date_col:
                            d = ws.cell(row=row_idx, column=date_col).value
                            if isinstance(d, datetime):
                                row_dt = d
                            elif isinstance(d, date):
                                row_dt = datetime.combine(d, time(0, 0, 0))
                        
                        if not row_dt:
                            continue
                        count += 1
                        if first_dt is None or row_dt < first_dt:
                            first_dt = row_dt
                        if last_dt is None or row_dt > last_dt:
                            last_dt = row_dt
                    except:
                        continue
                
                return count, first_dt, last_dt
            
            # 解析安卓sheet
            for sheet_name in android_sheets:
                ws = wb[sheet_name]
                cnt, first, last = _parse_sheet(ws)
                result['android_count'] += cnt
                if first:
                    if result['android_first'] is None or first < result['android_first']:
                        result['android_first'] = first
                if last:
                    if result['android_last'] is None or last > result['android_last']:
                        result['android_last'] = last
            
            # 解析苹果sheet
            for sheet_name in ios_sheets:
                ws = wb[sheet_name]
                cnt, first, last = _parse_sheet(ws)
                result['ios_count'] += cnt
                if first:
                    if result['ios_first'] is None or first < result['ios_first']:
                        result['ios_first'] = first
                if last:
                    if result['ios_last'] is None or last > result['ios_last']:
                        result['ios_last'] = last
            
            wb.close()
            
            result['total_count'] = result['android_count'] + result['ios_count']
            result['today_count'] = result['total_count']
            
            if result['ios_count'] > 0 and result['android_count'] == 0:
                result['work_type'] = '苹果'
            elif result['ios_count'] > 0 and result['android_count'] > 0:
                result['work_type'] = '安卓+苹果'
            
            all_firsts = [t for t in [result['android_first'], result['ios_first']] if t is not None]
            all_lasts = [t for t in [result['android_last'], result['ios_last']] if t is not None]
            if all_firsts:
                result['overall_first'] = min(all_firsts)
            if all_lasts:
                result['overall_last'] = max(all_lasts)
            
            result['android_hours'] = self._calc_app_work_hours(
                result['android_first'], result['android_last'])
            result['ios_hours'] = self._calc_app_work_hours(
                result['ios_first'], result['ios_last'])
            result['work_hours'] = result['android_hours'] + result['ios_hours']
            
            if result['work_hours'] > 0 and result['total_count'] > 0:
                result['upph'] = round(result['total_count'] / result['work_hours'], 2)
        
        except Exception as e:
            print(f"[parse_app_xlsx] 解析失败: {e}")
            import traceback
            traceback.print_exc()
        
        return result

    def get_group_efficiency(self, target_date: str = None) -> Dict:
        """
        获取所有组别的效率汇总
        """
        if target_date is None:
            target_date = date.today().strftime('%Y-%m-%d')

        cache_key = self._get_cache_key('group_efficiency', target_date)
        cached = _cache.get(cache_key)
        if cached is not None:
            return cached

        photo_data = self.get_photo_efficiency(target_date)
        auto_data = self.get_automation_efficiency(target_date)
        app_data = self.get_app_efficiency(target_date)

        # 合并所有组别
        all_groups = {}

        for g in photo_data.get('groups', []):
            name = g['group_name']
            if name not in all_groups:
                all_groups[name] = {
                    'group_name': name,
                    'photo': {'total_count': 0, 'staff_count': 0, 'achievement_rate': 0, 'avg_efficiency': 0},
                    'automation': {'total_count': 0, 'staff_count': 0, 'achievement_rate': 0, 'avg_efficiency': 0},
                    'app': {'total_count': 0, 'staff_count': 0, 'achievement_rate': 0, 'avg_efficiency': 0},
                }
            all_groups[name]['photo'] = {
                'total_count': g['total_count'],
                'staff_count': g['staff_count'],
                'achievement_rate': g['achievement_rate'],
                'avg_efficiency': g['avg_efficiency'],
            }

        for g in auto_data.get('groups', []):
            name = g['group_name']
            if name not in all_groups:
                all_groups[name] = {
                    'group_name': name,
                    'photo': {'total_count': 0, 'staff_count': 0, 'achievement_rate': 0, 'avg_efficiency': 0},
                    'automation': {'total_count': 0, 'staff_count': 0, 'achievement_rate': 0, 'avg_efficiency': 0},
                    'app': {'total_count': 0, 'staff_count': 0, 'achievement_rate': 0, 'avg_efficiency': 0},
                }
            all_groups[name]['automation'] = {
                'total_count': g['total_count'],
                'staff_count': g['staff_count'],
                'achievement_rate': g['achievement_rate'],
                'avg_efficiency': g['avg_efficiency'],
            }

        result = {
            'date': target_date,
            'groups': list(all_groups.values()),
            'total_staff': sum(g.get('photo', {}).get('staff_count', 0) + g.get('automation', {}).get('staff_count', 0) + g.get('app', {}).get('staff_count', 0) for g in all_groups.values()),
            'total_output': sum(g.get('photo', {}).get('total_count', 0) + g.get('automation', {}).get('total_count', 0) + g.get('app', {}).get('total_count', 0) for g in all_groups.values()),
            'overall_achievement': round(sum(g.get('photo', {}).get('achievement_rate', 0) + g.get('automation', {}).get('achievement_rate', 0) for g in all_groups.values()) / max(len(all_groups)*2, 1), 1),
            'avg_efficiency': 0,
            'all_staff': [],
            'photo_summary': {
                'total_count': photo_data['total_count'],
                'staff_count': photo_data['staff_count'],
                'avg_efficiency': photo_data['avg_efficiency'],
                'achievement_rate': photo_data['achievement_rate'],
            },
            'automation_summary': {
                'total_count': auto_data['total_count'],
                'staff_count': auto_data['staff_count'],
                'avg_efficiency': auto_data['avg_efficiency'],
                'achievement_rate': auto_data['achievement_rate'],
            },
            'app_summary': {
                'total_count': app_data.get('total_count', 0),
                'staff_count': app_data.get('staff_count', 0),
                'avg_efficiency': app_data.get('avg_efficiency', 0),
                'achievement_rate': app_data.get('achievement_rate', 0),
            },
        }

        _cache.set(cache_key, result)
        return result

    def get_home_kpi(self) -> Dict:
        """
        首页KPI汇总
        包含今日、近7天、近30天的关键指标
        """
        cache_key = 'home_kpi'
        cached = _cache.get(cache_key)
        if cached is not None:
            return cached

        today = date.today()
        today_str = today.strftime('%Y-%m-%d')

        # 今日数据
        photo_today = self.get_photo_efficiency(today_str)
        auto_today = self.get_automation_efficiency(today_str)

        # 近7天
        seven_days_ago = (today - timedelta(days=6)).strftime('%Y-%m-%d')
        photo_7d = self._get_period_photo_efficiency(seven_days_ago, today_str)
        auto_7d = self._get_period_automation_efficiency(seven_days_ago, today_str)

        # 近30天
        thirty_days_ago = (today - timedelta(days=29)).strftime('%Y-%m-%d')
        photo_30d = self._get_period_photo_efficiency(thirty_days_ago, today_str)
        auto_30d = self._get_period_automation_efficiency(thirty_days_ago, today_str)

        result = {
            'today': {
                'date': today_str,
                'photo': {
                    'total_count': photo_today['total_count'],
                    'staff_count': photo_today['staff_count'],
                    'avg_efficiency': photo_today['avg_efficiency'],
                    'achievement_rate': photo_today['achievement_rate'],
                },
                'automation': {
                    'total_count': auto_today['total_count'],
                    'staff_count': auto_today['staff_count'],
                    'avg_efficiency': auto_today['avg_efficiency'],
                    'achievement_rate': auto_today['achievement_rate'],
                },
            },
            'last_7_days': {
                'photo': photo_7d,
                'automation': auto_7d,
            },
            'last_30_days': {
                'photo': photo_30d,
                'automation': auto_30d,
            },
            'session_expired': {
                'godzilla': self.godzilla_client.session_expired,
                'mirror': self.mirror_client.session_expired,
            },
        }

        _cache.set(cache_key, result)
        return result

    def _get_period_photo_efficiency(self, start_date: str, end_date: str) -> Dict:
        """获取一段时间内的拍照效率汇总（按天计算后汇总）"""
        cache_key = self._get_cache_key('period_photo', start_date, end_date)
        cached = _cache.get(cache_key)
        if cached is not None:
            return cached

        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()

        # 按天分别计算各员工的工作时长和产量，再汇总
        daily_emp_stats = {}  # emp_no -> {total_count, work_hours, target_efficiency}
        current = start
        while current <= end:
            date_str = current.strftime('%Y-%m-%d')
            records = self.get_godzilla_records(date_str)
            day_staff = self._aggregate_by_staff(records, 'godzilla')
            for emp_no, s in day_staff.items():
                if emp_no not in daily_emp_stats:
                    daily_emp_stats[emp_no] = {
                        'total_count': 0,
                        'work_hours': 0,
                        'target_efficiency': s.get('target_efficiency', 0),
                        'achievement_days': 0,
                        'total_days': 0,
                    }
                daily_emp_stats[emp_no]['total_count'] += s['total_count']
                daily_emp_stats[emp_no]['work_hours'] += s['work_hours']
                if s['work_hours'] > 0:
                    daily_emp_stats[emp_no]['total_days'] += 1
                    if s['achievement_rate'] >= 100:
                        daily_emp_stats[emp_no]['achievement_days'] += 1
            current += timedelta(days=1)

        # 计算总体效率
        total_count = sum(s['total_count'] for s in daily_emp_stats.values())
        total_hours = sum(s['work_hours'] for s in daily_emp_stats.values())
        avg_eff = round(total_count / total_hours, 1) if total_hours > 0 else 0

        # 计算平均达标率（按员工平均）
        rates = []
        for emp_no, s in daily_emp_stats.items():
            if s['work_hours'] > 0 and s['target_efficiency'] > 0:
                emp_eff = s['total_count'] / s['work_hours']
                rates.append(round(emp_eff / s['target_efficiency'] * 100, 1))
        avg_rate = round(sum(rates) / len(rates), 1) if rates else 0

        result = {
            'start_date': start_date,
            'end_date': end_date,
            'total_count': total_count,
            'staff_count': len(daily_emp_stats),
            'avg_efficiency': avg_eff,
            'achievement_rate': avg_rate,
            'total_work_hours': round(total_hours, 1),
        }

        _cache.set(cache_key, result)
        return result

    def _get_period_automation_efficiency(self, start_date: str, end_date: str) -> Dict:
        """获取一段时间内的自动化效率汇总（按天计算后汇总）"""
        cache_key = self._get_cache_key('period_auto', start_date, end_date)
        cached = _cache.get(cache_key)
        if cached is not None:
            return cached

        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()

        # 按天分别计算各员工的工作时长和产量，再汇总
        daily_emp_stats = {}
        current = start
        while current <= end:
            date_str = current.strftime('%Y-%m-%d')
            records = self.get_mirror_records(date_str)
            day_staff = self._aggregate_by_staff(records, 'mirror')
            for emp_no, s in day_staff.items():
                if emp_no not in daily_emp_stats:
                    daily_emp_stats[emp_no] = {
                        'total_count': 0,
                        'work_hours': 0,
                        'target_efficiency': s.get('target_efficiency', 0),
                        'achievement_days': 0,
                        'total_days': 0,
                    }
                daily_emp_stats[emp_no]['total_count'] += s['total_count']
                daily_emp_stats[emp_no]['work_hours'] += s['work_hours']
                if s['work_hours'] > 0:
                    daily_emp_stats[emp_no]['total_days'] += 1
                    if s['achievement_rate'] >= 100:
                        daily_emp_stats[emp_no]['achievement_days'] += 1
            current += timedelta(days=1)

        total_count = sum(s['total_count'] for s in daily_emp_stats.values())
        total_hours = sum(s['work_hours'] for s in daily_emp_stats.values())
        avg_eff = round(total_count / total_hours, 1) if total_hours > 0 else 0

        rates = []
        for emp_no, s in daily_emp_stats.items():
            if s['work_hours'] > 0 and s['target_efficiency'] > 0:
                emp_eff = s['total_count'] / s['work_hours']
                rates.append(round(emp_eff / s['target_efficiency'] * 100, 1))
        avg_rate = round(sum(rates) / len(rates), 1) if rates else 0

        result = {
            'start_date': start_date,
            'end_date': end_date,
            'total_count': total_count,
            'staff_count': len(daily_emp_stats),
            'avg_efficiency': avg_eff,
            'achievement_rate': avg_rate,
            'total_work_hours': round(total_hours, 1),
        }

        _cache.set(cache_key, result)
        return result


# 全局服务实例
_efficiency_service: Optional[EfficiencyService] = None


def get_efficiency_service() -> EfficiencyService:
    """获取全局效率服务实例"""
    global _efficiency_service
    if _efficiency_service is None:
        _efficiency_service = EfficiencyService()
    return _efficiency_service
