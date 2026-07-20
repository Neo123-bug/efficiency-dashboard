# -*- coding: utf-8 -*-
"""
效率计算服务 - 从三个数据源拉数据，统一计算效率
- 按人员汇总
- 按日期计算（今日、近7天、近30天）
- 按组别分类
- 计算达标率
- 带缓存机制
"""
import json
import time
import logging
from datetime import datetime, date, timedelta
from typing import Dict, List, Optional, Tuple

from services.data_sources import GodzillaClient, MirrorClient, WatcherClient, GuangheClient, XRayClient
from config import EFFICIENCY_CONFIG, STAFF_ROSTER_PATH, CACHE_TTL_SECONDS

logger = logging.getLogger(__name__)


# ==================== 员工花名册 ====================

def load_staff_roster() -> Dict[str, Dict]:
    """
    加载员工花名册，返回 {employee_no: staff_info} 的字典
    """
    try:
        with open(STAFF_ROSTER_PATH, 'r', encoding='utf-8') as f:
            data = json.load(f)
        roster = {}
        for s in data.get('staff', []):
            emp_no = str(s.get('employee_no', ''))
            if not emp_no:
                emp_no = 'NAME_' + s.get('employee_name', 'unknown')
            roster[emp_no] = s
        return roster
    except Exception as e:
        logger.error(f'Load staff roster failed: {e}')
        return {}


def get_staff_info(employee_no: str, roster: Dict[str, Dict] = None) -> Dict:
    """获取员工信息"""
    if roster is None:
        roster = load_staff_roster()
    return roster.get(str(employee_no), {})


def get_staff_group(employee_no: str, roster: Dict[str, Dict] = None) -> str:
    """获取员工所属组别"""
    info = get_staff_info(employee_no, roster)
    return info.get('dept_group', '未分组')


def get_staff_name(employee_no: str, roster: Dict[str, Dict] = None) -> str:
    """获取员工姓名"""
    info = get_staff_info(employee_no, roster)
    return info.get('employee_name', employee_no)


def _is_after_lunch_cutoff(dt) -> bool:
    """判断某时间点是否晚于 13:30（用于工时剔除1小时判定）。"""
    if dt is None:
        return False
    return dt.hour > 13 or (dt.hour == 13 and dt.minute >= 30)


def _span_hours(first, last) -> float:
    """计算首末台时间跨度（小时），若最后一台晚于 13:30 则剔除1小时。无有效跨度返回 0。"""
    if not first or not last or last <= first:
        return 0.0
    dur_sec = (last - first).total_seconds()
    if _is_after_lunch_cutoff(last):
        dur_sec = max(dur_sec - 3600, 0)
    return round(dur_sec / 3600, 2)


def _spread_hourly(m: Dict) -> None:
    """将 total_count 按 first_start→last_end 的工作时段均匀分摊到 hourly_counts。
    若最后一台晚于 13:30，午餐时段(13点)不计入。"""
    first = m.get('first_start')
    last = m.get('last_end')
    count = m.get('total_count', 0)
    if not first or not last or count <= 0:
        return
    first_h = first.hour
    last_h = last.hour
    has_lunch = _is_after_lunch_cutoff(last)
    total_hours = last_h - first_h + 1
    if has_lunch and first_h < 13 <= last_h:
        total_hours -= 1
    if total_hours <= 0:
        return
    per_hour = count / total_hours
    if 'hourly_counts' not in m or not isinstance(m.get('hourly_counts'), dict):
        m['hourly_counts'] = {}
    for h in range(first_h, last_h + 1):
        if has_lunch and h == 13:
            continue
        m['hourly_counts'][h] = round(m['hourly_counts'].get(h, 0) + per_hour)


def get_target_efficiency(employee_no: str, data_source: str,
                          roster: Dict[str, Dict] = None) -> float:
    """
    获取员工的目标效率（UPPH标准）
    根据员工类型（全职/兼职）和数据源不同，标准不同
    """
    info = get_staff_info(employee_no, roster)
    emp_type = info.get('employee_type', '')

    if data_source == 'godzilla' or data_source.startswith('watcher_godzilla'):
        if '全职' in emp_type:
            return EFFICIENCY_CONFIG['fulltime_standard']
        else:
            return EFFICIENCY_CONFIG['parttime_standard']
    elif data_source == 'mirror':
        if '全职' in emp_type:
            return EFFICIENCY_CONFIG['mirror_fulltime_standard']
        else:
            return EFFICIENCY_CONFIG['mirror_parttime_standard']
    else:
        return 100.0


# ==================== 缓存机制 ====================

class CacheStore:
    """简单的内存缓存"""

    def __init__(self, ttl: int = CACHE_TTL_SECONDS):
        self._cache: Dict[str, Tuple[float, any]] = {}
        self.ttl = ttl

    def get(self, key: str) -> Optional[any]:
        """获取缓存，过期返回None"""
        if key not in self._cache:
            return None
        timestamp, value = self._cache[key]
        if time.time() - timestamp > self.ttl:
            del self._cache[key]
            return None
        return value

    def set(self, key: str, value: any):
        """设置缓存"""
        self._cache[key] = (time.time(), value)

    def invalidate(self, prefix: str = None):
        """清除缓存"""
        if prefix:
            keys_to_del = [k for k in self._cache if k.startswith(prefix)]
            for k in keys_to_del:
                del self._cache[k]
        else:
            self._cache.clear()


# 全局缓存实例
_cache = CacheStore()


def invalidate_cache(prefix: str = None):
    """清除缓存"""
    _cache.invalidate(prefix)


# ==================== 效率计算服务 ====================

def calc_photo_target(s, is_fulltime=None):
    """拍照人员综合目标效率（按处理量加权混合）：
        - 哥斯拉各服务标准：PHOTO_SERVICE_STANDARDS（PJTPZ=245，其余全职130/兼职120）
        - 光合拍照：按 GUANGHE_CATEGORY_STANDARDS 各品类标准（笔记本10.7/平板电脑27.4/...）
        - 拍照魔方（PJTPZ，watcher）：统一按 PJTPZ 标准
        - 瑕疵图（position_tag='瑕疵图'）：FLAW_PHOTO_STANDARDS（C2B=63.6，拍拍=36.7）
    拍拍(PPPZ) = 哥斯拉拍拍(130/120) + 光合拍照(各品类标准) 的混合加权。"""
    from config import (PHOTO_SERVICE_STANDARDS, GUANGHE_CATEGORY_STANDARDS,
                       GUANGHE_CATEGORY_ID_TO_NAME, GUANGHE_DEFAULT_STANDARD,
                       FLAW_PHOTO_STANDARDS)
    if is_fulltime is None:
        emp_type = s.get('employee_type', '全职') or '全职'
        is_fulltime = '全职' in emp_type

    def _std_for(svc):
        info = PHOTO_SERVICE_STANDARDS.get(svc, {})
        if isinstance(info, dict):
            return info.get('fulltime_standard', 130.0) if is_fulltime else info.get('parttime_standard', 120.0)
        return float(info) if info else 130.0

    # 瑕疵图人员优先使用瑕疵图标准
    if s.get('position_tag', '') == '瑕疵图':
        svc_counts = dict(s.get('service_standard_counts', {}))
        cube_cnt = s.get('photo_cube_count', 0)
        if isinstance(cube_cnt, (int, float)) and cube_cnt > 0:
            svc_counts['PJTPZ'] = svc_counts.get('PJTPZ', 0) + cube_cnt
        gh_cats = s.get('guanghe_categories', {}) or {}
        gh_total = sum(v.get('count', 0) if isinstance(v, dict) else int(v)
                      for v in gh_cats.values())
        if gh_total > 0:
            svc_counts['PPPZ'] = svc_counts.get('PPPZ', 0) + gh_total
        c2b_keys = {'ZJ0074', 'C2B'}
        c2b_cnt = sum(v for k, v in svc_counts.items()
                      if k in c2b_keys and isinstance(v, (int, float)))
        pp_cnt = sum(v for k, v in svc_counts.items()
                     if k not in c2b_keys and isinstance(v, (int, float)))
        if c2b_cnt > 0 and pp_cnt > 0:
            return round((c2b_cnt * FLAW_PHOTO_STANDARDS['C2B'] + pp_cnt * FLAW_PHOTO_STANDARDS['拍拍'])
                             / (c2b_cnt + pp_cnt), 1)
        elif c2b_cnt > 0:
            return FLAW_PHOTO_STANDARDS['C2B']
        else:
            return FLAW_PHOTO_STANDARDS['拍拍']

    parts = []  # [(count, target_eff)]
    # 1. 哥斯拉各服务标准
    svc_counts = s.get('service_standard_counts', {}) or {}
    for svc, cnt in svc_counts.items():
        cnt_val = cnt.get('count', 0) if isinstance(cnt, dict) else cnt
        if not isinstance(cnt_val, (int, float)) or cnt_val <= 0:
            continue
        parts.append((cnt_val, _std_for(svc)))
    # 2. 光合拍照：按品类标准
    gh_cats = s.get('guanghe_categories', {}) or {}
    for cid, info in gh_cats.items():
        cnt = info.get('count', 0) if isinstance(info, dict) else info
        if not isinstance(cnt, (int, float)) or cnt <= 0:
            continue
        cid_int = int(cid) if str(cid).isdigit() else cid
        cat_name = GUANGHE_CATEGORY_ID_TO_NAME.get(cid_int)
        target = GUANGHE_CATEGORY_STANDARDS.get(cat_name, GUANGHE_DEFAULT_STANDARD) if cat_name else GUANGHE_DEFAULT_STANDARD
        parts.append((cnt, target))
    # 3. 拍照魔方（PJTPZ，watcher）：统一按 PJTPZ 标准
    cube_cnt = s.get('photo_cube_count', 0)
    if isinstance(cube_cnt, (int, float)) and cube_cnt > 0:
        cube_std = PHOTO_SERVICE_STANDARDS.get('PJTPZ', {})
        cube_target = (cube_std.get('fulltime_standard', 245.0) if is_fulltime
                      else cube_std.get('parttime_standard', 245.0))
        parts.append((cube_cnt, cube_target))

    if not parts:
        return 130.0
    total_cnt = sum(p[0] for p in parts)
    if total_cnt <= 0:
        return 130.0
    return round(sum(p[0] * p[1] for p in parts) / total_cnt, 1)


class EfficiencyService:
    """
    统一的效率计算服务
    从哥斯拉、魔镜、Watcher三个数据源拉取数据，
    进行汇总、计算UPPH、达标率等指标
    """

    def __init__(self):
        self.godzilla_client = GodzillaClient()
        self.mirror_client = MirrorClient()
        self.watcher_client = WatcherClient()
        self.guanghe_client = None  # 懒加载
        self.xray_client = None  # 懒加载
        self.roster = load_staff_roster()

    def reload_roster(self):
        """重新加载花名册，确保人员管理的改动实时生效"""
        self.roster = load_staff_roster()

    def _get_cache_key(self, prefix: str, *args) -> str:
        """生成缓存key"""
        return f'{prefix}:' + ':'.join(str(a) for a in args)

    # ---------- 数据获取（带缓存） ----------

    def get_godzilla_records(self, target_date: str) -> List[Dict]:
        """获取哥斯拉当日全部记录（带缓存，不过滤类目）

        注意：此方法返回原始全量数据，供拍照页和自动化页共同使用。
        拍照页如需只统计拍照类目，应在调用后自行过滤。
        """
        cache_key = self._get_cache_key('godzilla_records', target_date)
        cached = _cache.get(cache_key)
        if cached is not None:
            return cached

        records = self.godzilla_client.fetch_by_date(target_date)
        _cache.set(cache_key, records)
        return records

    def get_mirror_records(self, target_date: str) -> List[Dict]:
        """获取魔镜当日记录（带缓存）"""
        cache_key = self._get_cache_key('mirror_records', target_date)
        cached = _cache.get(cache_key)
        if cached is not None:
            return cached

        records = self.mirror_client.fetch_by_date(target_date)
        _cache.set(cache_key, records)
        return records

    # ---------- 按员工汇总 ----------

    def _aggregate_by_staff(self, records: List[Dict], data_source: str) -> Dict[str, Dict]:
        """
        将明细记录按员工汇总
        返回: {employee_no: {employee_no, employee_name, total_count,
                             total_duration_sec, work_hours, efficiency, ...}}
        """
        staff_map: Dict[str, Dict] = {}

        for rec in records:
            emp_no = str(rec.get('employee_no', ''))
            if not emp_no or emp_no == 'None':
                continue

            if emp_no not in staff_map:
                name = rec.get('employee_name') or get_staff_name(emp_no, self.roster)
                group = get_staff_group(emp_no, self.roster)
                staff_map[emp_no] = {
                    'employee_no': emp_no,
                    'employee_name': name,
                    'group_name': group,
                    'total_count': 0,
                    'total_duration_sec': 0,
                    'first_time': None,
                    'last_time': None,
                    'data_source': data_source,
                    'position_tag': rec.get('position_tag', ''),
                    'machine_ids': set(),
                    'hourly_counts': {},
                }

            s = staff_map[emp_no]

            # 累加数量 - 每条记录代表1个产品/1次质检，UPPH按产品数计算
            # 同时记录photo_count用于参考展示
            s['total_count'] += 1  # 产品数/质检次数次数

            # 收集设备编号
            m_id = rec.get('machine_id', '') or ''
            if m_id:
                s['machine_ids'].add(str(m_id))
            if 'photo_count' in rec and rec['photo_count']:
                s['total_photos'] = s.get('total_photos', 0) + rec['photo_count']

            # 累加工时
            dur = rec.get('total_duration_sec', 0) or rec.get('duration_sec', 0) or 0
            s['total_duration_sec'] += int(dur)

            # 记录首次和末次时间（用于计算在岗时长）
            start_time = rec.get('start_time') or rec.get('date', '')
            if start_time:
                try:
                    dt = datetime.strptime(start_time[:19], '%Y-%m-%d %H:%M:%S')
                    if s['first_time'] is None or dt < s['first_time']:
                        s['first_time'] = dt
                    if s['last_time'] is None or dt > s['last_time']:
                        s['last_time'] = dt
                    # 时段统计
                    hour = dt.hour
                    s['hourly_counts'][hour] = s['hourly_counts'].get(hour, 0) + 1
                except (ValueError, IndexError):
                    pass

        # 计算工作时长和效率
        for emp_no, s in staff_map.items():
            # 方法1: 用总耗时长 / 3600
            duration_hours = s['total_duration_sec'] / 3600.0 if s['total_duration_sec'] > 0 else 0

            # ── 工作时长计算（按数据源区分规则）──
            # 哥斯拉：只用「首台→末台」时间跨度（用户确认口径），末台≥13:00扣1h午休
            # 魔镜/其他：取 max(有记录小时数, 跨度) 的较大值，避免跨度偏小虚高
            span_hours = 0.0
            if s['first_time'] and s['last_time']:
                span = s['last_time'] - s['first_time']
                span_hours = span.total_seconds() / 3600.0
                # 跨度口径的午休剔除
                lunch_hour = EFFICIENCY_CONFIG['lunch_break_hour']
                lunch_dur = EFFICIENCY_CONFIG['lunch_break_duration_hours']
                if s['last_time'].hour >= lunch_hour and s['first_time'].hour < lunch_hour + 1:
                    span_hours = max(0, span_hours - lunch_dur)
                s['span_hours'] = round(span_hours, 2)
            else:
                s['span_hours'] = 0

            if data_source == 'godzilla':
                # 哥斯拉：严格按首末台跨度算工时，不混入"有记录小时数"
                work_hours = span_hours
            else:
                # 魔镜等：取 max(有记录小时数扣午休, 跨度) 防止跨度偏小
                distinct_hours = len(s.get('hourly_counts', {}))
                last_h = s['last_time'].hour if s['last_time'] else 0
                first_h = s['first_time'].hour if s['first_time'] else 0
                lunch_dur_g = EFFICIENCY_CONFIG['lunch_break_duration_hours']
                lunch_deduct = lunch_dur_g if (last_h >= 13 and first_h < 14) else 0.0
                distinct_hours_adj = max(0.0, distinct_hours - lunch_deduct)
                s['distinct_hours'] = distinct_hours
                work_hours = max(distinct_hours_adj, span_hours)
            s['work_hours'] = round(work_hours, 2)
            s['duration_hours'] = round(duration_hours, 2)

            # 计算UPPH (Units Per Person Per Hour)
            if work_hours > 0:
                s['efficiency'] = round(s['total_count'] / work_hours, 1)
            else:
                s['efficiency'] = 0

            # 计算达标率
            target = get_target_efficiency(emp_no, data_source, self.roster)
            s['target_efficiency'] = target
            if target > 0 and s['efficiency'] > 0:
                s['achievement_rate'] = round(s['efficiency'] / target * 100, 1)
            else:
                s['achievement_rate'] = 0

            # 设备编号：去重后取前5个，用逗号分隔
            s['machine_count'] = len(s.get('machine_ids', set()))
            if s.get('machine_ids'):
                sorted_machines = sorted(s['machine_ids'])
                s['machine_id'] = ', '.join(sorted_machines[:5])
                if len(sorted_machines) > 5:
                    s['machine_id'] += f' 等{len(sorted_machines)}个'
            else:
                s['machine_id'] = ''
            # 删除set避免JSON序列化问题
            if 'machine_ids' in s:
                del s['machine_ids']

            # 转换datetime为字符串（JSON序列化需要）
            if s.get('first_time'):
                s['first_time'] = s['first_time'].strftime('%Y-%m-%d %H:%M:%S')
            if s.get('last_time'):
                s['last_time'] = s['last_time'].strftime('%Y-%m-%d %H:%M:%S')

        return staff_map

    # ---------- 按组别汇总 ----------

    def _aggregate_by_group(self, staff_map: Dict[str, Dict]) -> Dict[str, Dict]:
        """
        将员工数据按组别汇总
        """
        group_map: Dict[str, Dict] = {}

        for emp_no, staff in staff_map.items():
            group = staff['group_name']
            if group not in group_map:
                group_map[group] = {
                    'group_name': group,
                    'total_count': 0,
                    'total_work_hours': 0,
                    'staff_count': 0,
                    'achievement_rates': [],
                    'efficiencies': [],
                    'data_source': staff['data_source'],
                }

            g = group_map[group]
            g['total_count'] += staff['total_count']
            g['total_work_hours'] += staff['work_hours']
            g['staff_count'] += 1
            if staff['efficiency'] > 0:
                g['efficiencies'].append(staff['efficiency'])
            if staff['achievement_rate'] > 0:
                g['achievement_rates'].append(staff['achievement_rate'])

        # 计算组级别指标
        for group, g in group_map.items():
            if g['total_work_hours'] > 0:
                g['avg_efficiency'] = round(g['total_count'] / g['total_work_hours'], 1)
            else:
                g['avg_efficiency'] = 0

            if g['achievement_rates']:
                g['achievement_rate'] = round(
                    sum(g['achievement_rates']) / len(g['achievement_rates']), 1
                )
            else:
                g['achievement_rate'] = 0

            # 达标人数
            g['achieved_count'] = sum(1 for r in g['achievement_rates'] if r >= 100)
            g['total_work_hours'] = round(g['total_work_hours'], 1)

        return group_map

    # ---------- 对外API: 单天数据 ----------

    def _calc_photo_target_full(self, s) -> float:
        """
        计算拍照人员的综合目标效率（三部分互斥加权）：
        1. 哥斯拉拍照：按服务标准查 PHOTO_SERVICE_STANDARDS（区分全职/兼职）
        2. 光合拍照（PPPZ）：按品类加权
        3. 拍照魔方（PJTPZ）：统一标准 PJTPZ_PHOTO_CUBE_STANDARD
        4. 瑕疵图拍照（position_tag='瑕疵图'）：使用 FLAW_PHOTO_STANDARDS（C2B=63.6, 拍拍=36.7）
        """
        from config import (PHOTO_SERVICE_STANDARDS, PJTPZ_PHOTO_CUBE_STANDARD,
                           FLAW_PHOTO_STANDARDS)
        from config import (GUANGHE_CATEGORY_STANDARDS, GUANGHE_CATEGORY_ID_TO_NAME,
                           GUANGHE_DEFAULT_STANDARD)

        # ── 瑕疵图人员优先使用瑕疵图标准 ──
        pos_tag = s.get('position_tag', '')
        if pos_tag == '瑕疵图':
            svc_counts = s.get('service_standard_counts', {})
            # 判断主服务标准归属 C2B 还是 拍拍
            combined = dict(svc_counts)
            cube_cnt = s.get('photo_cube_count', 0)
            if cube_cnt > 0:
                combined['PJTPZ'] = combined.get('PJTPZ', 0) + cube_cnt
            gh_cats = s.get('guanghe_categories', {})
            gh_total = sum(
                v.get('count', 0) if isinstance(v, dict) else int(v)
                for v in gh_cats.values()
            ) if isinstance(gh_cats, dict) else 0
            if gh_total > 0:
                combined['PPPZ'] = combined.get('PPPZ', 0) + gh_total

            c2b_keys = {'ZJ0074', 'C2B'}
            c2b_cnt = sum(combined.get(k, 0) for k in c2b_keys if isinstance(combined.get(k), (int, float)))
            pp_cnt = sum(combined.get(k, 0) for k in combined if k not in c2b_keys and isinstance(combined.get(k), (int, float)))

            if c2b_cnt > 0 and pp_cnt > 0:
                # 混合来源，按量加权平均
                return round((c2b_cnt * FLAW_PHOTO_STANDARDS['C2B'] +
                             pp_cnt * FLAW_PHOTO_STANDARDS['拍拍']) / (c2b_cnt + pp_cnt), 1)
            elif c2b_cnt > 0:
                return FLAW_PHOTO_STANDARDS['C2B']
            else:
                return FLAW_PHOTO_STANDARDS['拍拍']

        # 统一走模块级函数 calc_photo_target（哥斯拉各服务标准 + 光合各品类标准 + 拍照魔方 按量加权）
        return calc_photo_target(s)

    def _get_guanghe_client(self):
        if self.guanghe_client is None:
            self.guanghe_client = GuangheClient()
        return self.guanghe_client

    def _get_xray_client(self):
        if self.xray_client is None:
            self.xray_client = XRayClient()
        return self.xray_client

    def get_photo_efficiency(self, target_date: str = None) -> Dict:
        """
        获取拍照效率（多数据源合并）
        - 哥斯拉系统：拍拍拍照(PPPZ)、拍拍门店拍照(PPMDPZ)、拍机堂拍照(PJTPZ)、C2B门店入哥斯拉
        - 光合系统(PPPZ)：拍拍拍照补充，按工号精准匹配
        - Watcher拍照魔方(PJTPZ)：拍机堂拍照补充
        工时统一按"第一台到最后一台，13:30后结束减1小时"计算
        """
        if target_date is None:
            target_date = date.today().strftime('%Y-%m-%d')

        cache_key = self._get_cache_key('photo_efficiency', target_date)
        cached = _cache.get(cache_key)
        if cached is not None:
            return cached

        # 每次刷新重新加载花名册，确保人员管理改动实时生效
        self.reload_roster()

        # _dt 别名（提到函数开头，确保哥斯拉数据为空、下方 for 循环内 import 未执行时，光合段也能使用）
        from datetime import datetime as _dt

        # ===== 数据源1: 哥斯拉拍照记录 =====
        godzilla_records = self.get_godzilla_records(target_date)
        # 拍照页只保留拍照类目，排除全托质检(QTZJ)等手机质检
        from config import GODZILLA_PHOTO_CATEGORIES
        godzilla_records = [
            r for r in godzilla_records
            if str(r.get('service_standard', '') or r.get('serviceStandard', '')).strip()
            in GODZILLA_PHOTO_CATEGORIES
        ]
        raw_staff = self._aggregate_by_staff(godzilla_records, 'godzilla')

        # 统计每个员工的服务标准分布
        svc_std_map = {}  # emp_no -> {service_std: count}
        for r in godzilla_records:
            emp_no = str(r.get('employee_no', ''))
            if not emp_no or emp_no == 'None':
                continue
            svc_std = r.get('service_standard', '') or r.get('serviceStandard', '') or 'ZJ0074'
            if emp_no not in svc_std_map:
                svc_std_map[emp_no] = {}
            if svc_std not in svc_std_map[emp_no]:
                svc_std_map[emp_no][svc_std] = 0
            svc_std_map[emp_no][svc_std] += 1

        # 转换格式 + 重算工时（第一台到最后一台，13:30后减1小时）
        godzilla_staff = {}
        for emp_no, s in raw_staff.items():
            first = s.get('first_time')
            last = s.get('last_time')
            # 确保是datetime对象
            from datetime import datetime as _dt
            if isinstance(first, str):
                try: first = _dt.strptime(first[:19], '%Y-%m-%d %H:%M:%S')
                except: first = None
            if isinstance(last, str):
                try: last = _dt.strptime(last[:19], '%Y-%m-%d %H:%M:%S')
                except: last = None
            dur_sec = 0
            if first and last and last > first:
                dur_sec = (last - first).total_seconds()
                # 13:30以后结束减1小时午休
                if last.hour > 13 or (last.hour == 13 and last.minute >= 30):
                    dur_sec = max(dur_sec - 3600, 0)
            godzilla_staff[emp_no] = {
                'employee_no': emp_no,
                'name': s.get('employee_name', s.get('name', emp_no)),
                'total_count': s.get('total_count', 0),
                'total_duration_sec': dur_sec,
                'work_hours': round(dur_sec / 3600, 2),
                'efficiency': round(s.get('total_count', 0) / (dur_sec / 3600), 1) if dur_sec > 0 else 0,
                'service_standard_counts': svc_std_map.get(emp_no, {}),
                'godzilla_count': s.get('total_count', 0),
                'godzilla_duration_sec': s.get('total_duration_sec', 0),
                'machine_id': s.get('machine_id', ''),
                'hourly_counts': s.get('hourly_counts', {}),
                'data_sources': ['godzilla'],
                'first_start': first,
                'last_end': last,
            }

        # ===== 数据源2: Watcher 拍照魔方（拍机堂拍照 PJTPZ） =====
        # watcher返回的是标准化list，只有employee_name没有工号，按姓名匹配花名册
        watcher_photo = {}  # key=employee_no, value=list of {name, count, duration_sec, photo_method, service_standard, category}
        try:
            w_rows = self.watcher_client.fetch_photo_detail(target_date, target_date)
            if w_rows:
                # 按 姓名+photo_method+service_standard 聚合
                name_method_agg = {}
                for row in w_rows:
                    name = str(row.get('employee_name', '')).strip()
                    method = str(row.get('photo_method', '')).strip()
                    svc_std = str(row.get('service_standard', '')).strip()
                    cat = str(row.get('category', '')).strip()
                    cnt = int(row.get('count', 0) or 0)
                    dur = int(row.get('duration_sec', 0) or 0)
                    op_dur = int(row.get('op_duration_sec', 0) or 0)
                    if not name or cnt <= 0:
                        continue
                    key = (name, method, svc_std)
                    if key not in name_method_agg:
                        name_method_agg[key] = {
                            'name': name,
                            'photo_method': method,
                            'service_standard': svc_std,
                            'category': cat,
                            'count': 0,
                            'duration_sec': 0,
                            'op_duration_sec': 0,
                            'first_time': None,
                            'last_time': None,
                        }
                    name_method_agg[key]['count'] += cnt
                    name_method_agg[key]['duration_sec'] += dur
                    name_method_agg[key]['op_duration_sec'] += op_dur
                    # 累计首末台完成时间（取该聚合组 min/max）
                    ft = row.get('first_time')
                    lt = row.get('last_time')
                    cur = name_method_agg[key]
                    if ft and (cur['first_time'] is None or ft < cur['first_time']):
                        cur['first_time'] = ft
                    if lt and (cur['last_time'] is None or lt > cur['last_time']):
                        cur['last_time'] = lt
                # 用花名册反向查工号
                for (name, method, svc_std), info in name_method_agg.items():
                    emp_no = None
                    for en, r_info in self.roster.items():
                        if r_info.get('name') == name:
                            emp_no = en
                            break
                    if emp_no:
                        if emp_no not in watcher_photo:
                            watcher_photo[emp_no] = []
                        watcher_photo[emp_no].append({
                            'employee_no': emp_no,
                            'name': name,
                            'count': info['count'],
                            'duration_sec': info.get('duration_sec', 0),
                            'op_duration_sec': info.get('op_duration_sec', 0),
                            'photo_method': info['photo_method'],
                            'service_standard': info['service_standard'],
                            'category': info['category'],
                            'first_time': info.get('first_time'),
                            'last_time': info.get('last_time'),
                        })
            total_watcher = sum(item['count'] for items in watcher_photo.values() for item in items)
            print(f"Watcher拍照明细: {len(watcher_photo)}人, {total_watcher}台")
        except Exception as e:
            print(f"Watcher拍照魔方获取失败: {e}")

        # ===== 数据源3: 光合拍照（PPPZ） =====
        guanghe_staff = {}
        try:
            gh_client = self._get_guanghe_client()
            gh_records = gh_client.fetch_photo_tasks(date_str=target_date)
            if gh_records:
                for r in gh_records:
                    emp_no = str(r.get('employeeNo', '') or r.get('employee_no', '')).strip()
                    if not emp_no:
                        continue
                    cat_id = r.get('productCategoryId')
                    start_raw = r.get('startPictureTime') or r.get('startTime', '')
                    end_raw = r.get('endPictureTime') or r.get('endTime', '')

                    start_dt = None
                    end_dt = None
                    # 时间戳（毫秒）转datetime
                    try:
                        if start_raw and isinstance(start_raw, (int, float)):
                            start_dt = _dt.fromtimestamp(start_raw / 1000.0)
                        elif start_raw and isinstance(start_raw, str) and start_raw.isdigit():
                            start_dt = _dt.fromtimestamp(int(start_raw) / 1000.0)
                        elif start_raw:
                            s_clean = str(start_raw).split('.')[0]
                            start_dt = _dt.strptime(s_clean, '%Y-%m-%d %H:%M:%S')
                    except (ValueError, TypeError, OSError):
                        pass
                    try:
                        if end_raw and isinstance(end_raw, (int, float)):
                            end_dt = _dt.fromtimestamp(end_raw / 1000.0)
                        elif end_raw and isinstance(end_raw, str) and end_raw.isdigit():
                            end_dt = _dt.fromtimestamp(int(end_raw) / 1000.0)
                        elif end_raw:
                            e_clean = str(end_raw).split('.')[0]
                            end_dt = _dt.strptime(e_clean, '%Y-%m-%d %H:%M:%S')
                    except (ValueError, TypeError, OSError):
                        pass

                    if emp_no not in guanghe_staff:
                        guanghe_staff[emp_no] = {
                            'employee_no': emp_no,
                            'name': r.get('employeeName', '') or r.get('employee_name', ''),
                            'total_count': 0,
                            'first_start': start_dt,
                            'last_end': end_dt,
                            'categories': {},
                            'machine_ids': set(),
                        }

                    g = guanghe_staff[emp_no]
                    g['total_count'] += 1  # 1条记录=1台设备
                    cat_key = str(cat_id) if cat_id else 'unknown'
                    if cat_key not in g['categories']:
                        g['categories'][cat_key] = 0
                    g['categories'][cat_key] += 1

                    # 收集设备编号（光合: machineNo）
                    gh_device = (r.get('machineNo') or r.get('deviceCode')
                                 or r.get('deviceNo') or r.get('equipmentId') or '')
                    if gh_device:
                        g['machine_ids'].add(str(gh_device).strip())

                    # 时段统计
                    if start_dt:
                        h = start_dt.hour
                        g['hourly_counts'] = g.get('hourly_counts', {})
                        g['hourly_counts'][h] = g['hourly_counts'].get(h, 0) + 1

                    if start_dt and (g['first_start'] is None or start_dt < g['first_start']):
                        g['first_start'] = start_dt
                    if end_dt and (g['last_end'] is None or end_dt > g['last_end']):
                        g['last_end'] = end_dt

                # 计算光合工时（第一台到最后一台，13:30后减1小时）
                for emp_no, g in guanghe_staff.items():
                    first = g['first_start']
                    last = g['last_end']
                    dur_sec = 0
                    if first and last and last > first:
                        dur_sec = (last - first).total_seconds()
                        if last.hour > 13 or (last.hour == 13 and last.minute >= 30):
                            dur_sec = max(dur_sec - 3600, 0)
                    g['total_duration_sec'] = dur_sec
                    g['work_hours'] = round(dur_sec / 3600, 2)
                    # 设备编号：去重后取前5个，逗号分隔
                    if g.get('machine_ids'):
                        sorted_machines = sorted(g['machine_ids'])
                        g['machine_id'] = ', '.join(sorted_machines[:5])
                        if len(sorted_machines) > 5:
                            g['machine_id'] += f' 等{len(sorted_machines)}个'
                    else:
                        g['machine_id'] = ''
                    # 删除set避免序列化问题
                    if 'machine_ids' in g:
                        del g['machine_ids']

            print(f"光合拍照: {len(guanghe_staff)}人, {sum(v['total_count'] for v in guanghe_staff.values())}台")
        except Exception as e:
            print(f"光合拍照获取失败: {e}")

        # ===== 合并所有数据源到同一员工 =====
        merged_staff = {}

        # 先加哥斯拉
        for emp_no, s in godzilla_staff.items():
            merged_staff[emp_no] = dict(s)
            merged_staff[emp_no]['guanghe_categories'] = {}
            merged_staff[emp_no]['photo_cube_count'] = 0

        # 合并Watcher拍照数据 → 按photo_method分类
        # 拍照魔方 → PJTPZ(photo_cube_count)
        # 哥斯拉/光合等其他方式：量并入对应数据源（如果已有），否则跳过
        pre_existing = set(merged_staff.keys())  # 已在哥斯拉/光合中出现过的人员
        for emp_no, w_items in watcher_photo.items():
            for w in w_items:
                method = w.get('photo_method', '')
                count = w['count']
                dur_sec = w.get('duration_sec', 0)
                op_dur_sec = w.get('op_duration_sec', 0)
                svc_std = w.get('service_standard', '')
                
                if method == '拍照魔方':
                    # 拍照魔方 → PJTPZ
                    if emp_no in merged_staff:
                        m = merged_staff[emp_no]
                        m['total_count'] += count
                        m['photo_cube_count'] = m.get('photo_cube_count', 0) + count
                        m['op_duration_sec'] = m.get('op_duration_sec', 0) + op_dur_sec
                        if 'watcher' not in m['data_sources']:
                            m['data_sources'].append('watcher')
                        # 拍照魔方时段统计：按工作时段均匀分摊
                        if count > 0 and m.get('first_start') and m.get('last_end'):
                            first_h = m['first_start'].hour
                            last_h = m['last_end'].hour
                            has_lunch = m['last_end'].hour > 13 or (m['last_end'].hour == 13 and m['last_end'].minute >= 30)
                            total_hours = last_h - first_h + 1
                            if has_lunch and first_h < 13 <= last_h:
                                total_hours -= 1
                            if total_hours > 0:
                                per_hour = count / total_hours
                                if 'hourly_counts' not in m:
                                    m['hourly_counts'] = {}
                                for h in range(first_h, last_h + 1):
                                    if has_lunch and h == 13:
                                        continue
                                    m['hourly_counts'][h] = round(m['hourly_counts'].get(h, 0) + per_hour)
                    else:
                        # 纯拍照魔方人员（没有哥斯拉/光合数据）
                        # 工时按「第一台→最后一台手机时间跨度」计算（>13:30 剔1小时），
                        # 而非 sum(步骤一时长)——后者是设备累计耗时，会把工时压到 0.2h 导致 UPPH 上千。
                        # 先累计各聚合组的首末台时间，循环结束后再统一算 span。
                        if emp_no not in merged_staff:
                            merged_staff[emp_no] = {
                                'employee_no': emp_no,
                                'name': w['name'],
                                'total_count': 0,
                                'total_duration_sec': 0,
                                'work_hours': 0,
                                'efficiency': 0,
                                'service_standard_counts': {},
                                'guanghe_categories': {},
                                'photo_cube_count': 0,
                                'watcher_duration_sec': 0,
                                'op_duration_sec': 0,
                                'godzilla_count': 0,
                                'data_sources': ['watcher'],
                                'first_start': None,
                                'last_end': None,
                                'hourly_counts': {},
                            }
                        m = merged_staff[emp_no]
                        m['total_count'] += count
                        m['photo_cube_count'] += count
                        m['op_duration_sec'] += op_dur_sec
                        m['watcher_duration_sec'] += op_dur_sec
                        # 累计首末台完成时间跨度（跨所有拍照魔方聚合组取全局 min/max）
                        ft = w.get('first_time')
                        lt = w.get('last_time')
                        if ft and (m['first_start'] is None or ft < m['first_start']):
                            m['first_start'] = ft
                        if lt and (m['last_end'] is None or lt > m['last_end']):
                            m['last_end'] = lt
                        if 'watcher' not in m['data_sources']:
                            m['data_sources'].append('watcher')
                elif method == '哥斯拉':
                    # watcher里的哥斯拉方式数据 → 和哥斯拉数据源的量是同一批，跳过（避免重复）
                    # 只补充service_standard分布
                    if emp_no in merged_staff:
                        m = merged_staff[emp_no]
                        if svc_std and 'service_standard_counts' in m:
                            # 不重复加数量，只补服务标准信息
                            pass
                # 其他方式（光合拍照、室检App等）跳过，由对应数据源负责

        # ===== 纯拍照魔方人员工时收尾：按首末台跨度计算（>13:30 剔1小时）=====
        # 这些人员没有哥斯拉/光合数据，无法用哥斯拉时段跨度，必须在此用 watcher 首末台时间。
        # 只处理「本次新增」的纯 watcher 人员（pre_existing 之外的人员）。
        for emp_no in (set(merged_staff.keys()) - pre_existing):
            m = merged_staff.get(emp_no)
            if not m or 'watcher' not in m.get('data_sources', []):
                continue
            # 若已被光合合并则跳过，交给光合段统一算
            if 'guanghe' in m.get('data_sources', []):
                continue
            first = m.get('first_start')
            last = m.get('last_end')
            if first and last and last > first:
                span = _span_hours(first, last)
                if span > 0:
                    m['total_duration_sec'] = int(span * 3600)
                    m['work_hours'] = span
                    _spread_hourly(m)
                else:
                    # 跨度异常（首末相等等）→ 退回标准工作日估算
                    m['work_hours'] = 7.0
                    m['total_duration_sec'] = int(7.0 * 3600)
            else:
                # 无有效首末台时间 → 退回标准工作日估算（8h-1h午休=7h）
                m['work_hours'] = 7.0
                m['total_duration_sec'] = int(7.0 * 3600)


        # 合并光合拍照 → PPPZ
        for emp_no, g in guanghe_staff.items():
            if emp_no in merged_staff:
                m = merged_staff[emp_no]
                # 合并处理量
                m['total_count'] += g['total_count']
                m['guanghe_duration_sec'] = g.get('total_duration_sec', 0)
                m['guanghe_categories'] = dict(g['categories'])

                # 合并时段统计
                gh_hourly = g.get('hourly_counts', {})
                if gh_hourly:
                    if 'hourly_counts' not in m:
                        m['hourly_counts'] = {}
                    for h, c in gh_hourly.items():
                        m['hourly_counts'][h] = m['hourly_counts'].get(h, 0) + c
                # 注意：光合数据不计入service_standard_counts（避免和光合品类目标计算重复）
                # 光合品类单独存 guanghe_categories
                if 'guanghe' not in m['data_sources']:
                    m['data_sources'].append('guanghe')
                # 合并时间范围（取更早开始和更晚结束）
                if g['first_start']:
                    if m.get('first_start') is None or g['first_start'] < m['first_start']:
                        m['first_start'] = g['first_start']
                if g['last_end']:
                    if m.get('last_end') is None or g['last_end'] > m['last_end']:
                        m['last_end'] = g['last_end']
                # 重新计算总工时
                first = m.get('first_start')
                last = m.get('last_end')
                dur_sec = 0
                if first and last and last > first:
                    dur_sec = (last - first).total_seconds()
                    if last.hour > 13 or (last.hour == 13 and last.minute >= 30):
                        dur_sec = max(dur_sec - 3600, 0)
                m['total_duration_sec'] = dur_sec
                m['work_hours'] = round(dur_sec / 3600, 2)
                # 合并设备编号（哥斯拉+光合设备编号拼接）
                gh_mid = g.get('machine_id', '')
                if gh_mid:
                    existing_mid = m.get('machine_id', '')
                    if existing_mid and gh_mid not in existing_mid:
                        m['machine_id'] = existing_mid + ', ' + gh_mid
                    elif not existing_mid:
                        m['machine_id'] = gh_mid
            else:
                merged_staff[emp_no] = {
                    'employee_no': emp_no,
                    'name': g['name'],
                    'total_count': g['total_count'],
                    'total_duration_sec': g['total_duration_sec'],
                    'work_hours': g['work_hours'],
                    'efficiency': 0,
                    'service_standard_counts': {},
                    'guanghe_categories': dict(g['categories']),
                    'guanghe_duration_sec': g.get('total_duration_sec', 0),
                    'photo_cube_count': 0,
                    'godzilla_count': 0,
                    'machine_id': g.get('machine_id', ''),
                    'data_sources': ['guanghe'],
                    'hourly_counts': dict(g.get('hourly_counts', {})),
                    'first_start': g['first_start'],
                    'last_end': g['last_end'],
                }

        # ===== 瑕疵图拍照完成量（widget 7757 团队总量，用于卡片）=====
        try:
            flaw_total = self.watcher_client.fetch_flaw_photo(target_date, target_date)
        except Exception as e:
            print(f"瑕疵图拍照获取失败: {e}")
            flaw_total = 0

        # 瑕疵图按人员明细（widget 3535 + 是否完成瑕疵拍照='是'，含姓名与服务标准）
        # 用于把瑕疵图量按姓名计入对应拍照组人员（瑕疵图可由拍照组任意人员拍摄，非仅「瑕疵图」岗）
        try:
            flaw_detail = self.watcher_client.fetch_widget_data('flaw_photo_detail', target_date, target_date) or []
        except Exception as e:
            print(f"瑕疵图明细获取失败: {e}")
            flaw_detail = []

        # ===== 按花名册「拍照组」筛选 + 补全姓名 =====
        # 只保留「人员管理」中属于拍照组的人员：
        #   花名册 group_name == '拍照' 或 岗位标签 position_tag 含「拍照」「瑕疵图」。
        # 哥斯拉组（position_tag='哥斯拉'，手机质检岗）不属于拍照组，
        # 即使当天在哥斯拉系统中做了拍照工序，其量也不计入拍照效率页的人员明细。
        photo_staff = {}
        for emp_no, s in merged_staff.items():
            info = get_staff_info(emp_no, self.roster)
            pos_tag = info.get('position_tag', '')
            grp_name = info.get('group_name', '') or info.get('dept_group', '')
            is_photo_group = (
                (pos_tag and ('拍照' in pos_tag or '瑕疵图' in pos_tag))
                or grp_name == '拍照'
            )
            if not is_photo_group:
                continue
            s['name'] = info.get('name', s.get('name', emp_no))
            s['employee_type'] = info.get('employee_type', '')
            s['position_tag'] = pos_tag if pos_tag else s.get('position_tag', '')
            s['group_name'] = info.get('group_name', '')
            # 工时兜底：首末时间算出来为0时，用各数据源时长之和估算
            if s['work_hours'] <= 0:
                total_dur = 0
                has_non_watcher = False
                if 'guanghe_duration_sec' in s:
                    total_dur += s['guanghe_duration_sec']
                    has_non_watcher = True
                # 哥斯拉total_duration_sec也兜底（单条时长累加）
                if 'godzilla_duration_sec' in s:
                    total_dur += s['godzilla_duration_sec']
                    has_non_watcher = True
                # watcher_duration_sec 仅在有其他数据源时辅助补充（纯拍照魔方时不用）
                # 原因：watcher聚合的 sum(拍照时长) 是所有设备拍照时长累计，不是员工实际工作时间
                if has_non_watcher and 'watcher_duration_sec' in s:
                    total_dur += s['watcher_duration_sec']
                if total_dur > 0:
                    s['work_hours'] = round(total_dur / 3600, 2)
                    s['total_duration_sec'] = total_dur
                elif s.get('photo_cube_count', 0) > 0 or s['total_count'] > 0:
                    # 纯拍照魔方人员且无任何有效工时 → 按标准工作日估算（8小时-1小时午休=7小时）
                    s['work_hours'] = 7.0
            # 重算效率
            s['efficiency'] = round(s['total_count'] / s['work_hours'], 1) if s['work_hours'] > 0 else 0
            # 综合目标效率
            s['target_efficiency'] = self._calc_photo_target_full(s)
            # 达标率
            if s['target_efficiency'] > 0 and s['efficiency'] > 0:
                s['achievement_rate'] = round(s['efficiency'] / s['target_efficiency'] * 100, 1)
            else:
                s['achievement_rate'] = 0
            # 主服务标准（取最多的那个）
            svc_counts = s.get('service_standard_counts', {})
            # 合并watcher拍照魔方(PJTPZ)和光合(PPPZ)的量，统一判断主服务标准
            combined_counts = dict(svc_counts)
            cube_cnt = s.get('photo_cube_count', 0)
            if cube_cnt > 0:
                combined_counts['PJTPZ'] = combined_counts.get('PJTPZ', 0) + cube_cnt
            gh_total = sum(s.get('guanghe_categories', {}).values()) if isinstance(s.get('guanghe_categories', {}), dict) else 0
            if gh_total > 0:
                combined_counts['PPPZ'] = combined_counts.get('PPPZ', 0) + gh_total
            if combined_counts:
                s['service_standard'] = max(combined_counts.items(), key=lambda x: x[1] if isinstance(x[1], (int, float)) else x[1].get('count', 0))[0]
            else:
                s['service_standard'] = 'ZJ0074'

            photo_staff[emp_no] = s

        # ===== 瑕疵图按姓名计入拍照组人员综合效率与处理量 =====
        # 瑕疵图可由拍照组任意人员拍摄（数据带「步骤一处理人」姓名），故按姓名匹配花名册，
        # 将每人瑕疵图量并入其 total_count，并按达标标准(FLAW_PHOTO_STANDARDS)计入综合效率。
        # 不再限于「瑕疵图」岗位人员（如欧紫月）一人，纠正此前只绑一人的错误归因。
        # 初始化瑕疵图字段，避免下游缺键
        for _s in photo_staff.values():
            _s.setdefault('flaw_count', 0)
            _s.setdefault('flaw_by_standard', {})

        if flaw_detail:
            # 姓名→工号映射（花名册）
            name_to_empno = {}
            for en, r in self.roster.items():
                nm = r.get('name')
                if nm:
                    name_to_empno.setdefault(nm, en)

            # 先按 (工号, 服务标准) 聚合瑕疵图量
            flaw_agg = {}  # (emp_no, svc) -> count
            for r in flaw_detail:
                if not isinstance(r, dict):
                    continue
                nm = str(r.get('步骤一处理人', '') or '').strip()
                if not nm:
                    continue
                svc = str(r.get('服务标准', '') or '').strip()
                try:
                    cnt = int(r.get('物品量', 0) or 0)
                except (ValueError, TypeError):
                    cnt = 0
                if cnt <= 0:
                    continue
                emp_no = name_to_empno.get(nm)
                if not emp_no:
                    continue
                # 仅计入拍照组人员
                info = get_staff_info(emp_no, self.roster)
                pos_tag = info.get('position_tag', '')
                grp = info.get('group_name', '') or info.get('dept_group', '')
                if not ((pos_tag and ('拍照' in pos_tag or '瑕疵图' in pos_tag)) or grp == '拍照'):
                    continue
                key = (emp_no, svc)
                flaw_agg[key] = flaw_agg.get(key, 0) + cnt

            # 并入对应人员，并重算效率/达标率
            for (emp_no, svc), cnt in flaw_agg.items():
                if emp_no in photo_staff:
                    s = photo_staff[emp_no]
                else:
                    # 当天仅拍了瑕疵图、无其他拍照数据的拍照组人员：新建条目（工时按标准班次估算）
                    info = get_staff_info(emp_no, self.roster)
                    s = {
                        'employee_no': emp_no,
                        'name': info.get('name', emp_no),
                        'employee_type': info.get('employee_type', ''),
                        'position_tag': info.get('position_tag', ''),
                        'group_name': info.get('group_name', ''),
                        'total_count': 0,
                        'total_duration_sec': 0,
                        'work_hours': 7.0,
                        'efficiency': 0,
                        'service_standard_counts': {},
                        'guanghe_categories': {},
                        'photo_cube_count': 0,
                        'godzilla_count': 0,
                        'flaw_count': 0,
                        'flaw_by_standard': {},
                        'data_sources': ['watcher'],
                        'first_start': None,
                        'last_end': None,
                        'hourly_counts': {},
                    }
                    photo_staff[emp_no] = s
                s['total_count'] = s.get('total_count', 0) + cnt
                s['flaw_count'] = s.get('flaw_count', 0) + cnt
                if 'flaw_by_standard' not in s:
                    s['flaw_by_standard'] = {}
                s['flaw_by_standard'][svc] = s['flaw_by_standard'].get(svc, 0) + cnt
                if 'watcher' not in s['data_sources']:
                    s['data_sources'].append('watcher')
                # 重算效率与达标率（工时沿用其当班时长，瑕疵图在同一班次内完成）
                s['efficiency'] = round(s['total_count'] / s['work_hours'], 1) if s.get('work_hours', 0) > 0 else 0
                s['target_efficiency'] = self._calc_photo_target_full(s)
                if s['target_efficiency'] > 0 and s['efficiency'] > 0:
                    s['achievement_rate'] = round(s['efficiency'] / s['target_efficiency'] * 100, 1)
                else:
                    s['achievement_rate'] = 0
                # 更新主服务标准（含瑕疵图量）
                svc_counts = s.get('service_standard_counts', {}) or {}
                combined = dict(svc_counts)
                cube_cnt = s.get('photo_cube_count', 0)
                if cube_cnt > 0:
                    combined['PJTPZ'] = combined.get('PJTPZ', 0) + cube_cnt
                gh_total = sum(v.get('count', 0) if isinstance(v, dict) else v for v in s.get('guanghe_categories', {}).values())
                if gh_total > 0:
                    combined['PPPZ'] = combined.get('PPPZ', 0) + gh_total
                for fsvc, fcnt in s.get('flaw_by_standard', {}).items():
                    combined[fsvc] = combined.get(fsvc, 0) + fcnt
                if combined:
                    s['service_standard'] = max(combined.items(), key=lambda x: x[1] if isinstance(x[1], (int, float)) else x[1].get('count', 0))[0]
                print(f"瑕疵图计入 {s.get('name')}: +{cnt}台(标准{svc}) → total_count={s['total_count']} eff={s['efficiency']} 达成={s['achievement_rate']}%")

        # ===== 服务标准分组 =====
        # 按实际服务标准统计（一个人可以在多个组里）
        # 合并所有数据源的服务标准量：哥斯拉svc_counts + 拍照魔方(PJTPZ) + 光合(PPPZ)
        svc_groups = {}
        svc_staff_sets = {}  # 单独用dict存人员集合，避免set进结果
        for emp_no, s in photo_staff.items():
            # 计算各服务标准的量
            svc_counts = dict(s.get('service_standard_counts', {}) or {})
            # 拍照魔方 → PJTPZ
            cube_cnt = s.get('photo_cube_count', 0)
            if cube_cnt > 0:
                svc_counts['PJTPZ'] = svc_counts.get('PJTPZ', 0) + cube_cnt
            # 光合 → PPPZ
            gh_cats = s.get('guanghe_categories', {}) or {}
            gh_total = sum(v.get('count', 0) if isinstance(v, dict) else v for v in gh_cats.values())
            if gh_total > 0:
                svc_counts['PPPZ'] = svc_counts.get('PPPZ', 0) + gh_total
            
            work_h = s.get('work_hours', 0)
            eff = s.get('efficiency', 0)
            ach_rate = s.get('achievement_rate', 0)
            is_achieved = ach_rate >= 100
            
            for svc, count in svc_counts.items():
                if count <= 0:
                    continue
                # 只统计4个标准服务
                if svc not in ('PPPZ', 'PPMDPZ', 'PJTPZ', 'ZJ0074'):
                    continue
                if svc not in svc_groups:
                    svc_groups[svc] = {
                        'group_name': svc,
                        'total_count': 0,
                        'staff_count': 0,
                        'total_work_hours': 0,
                        'efficiencies': [],
                        'achievement_rates': [],
                        'achieved_count': 0,
                    }
                    svc_staff_sets[svc] = set()
                g = svc_groups[svc]
                g['total_count'] += count
                # 人员按独立人头算（一个人在一个标准里只算1次）
                if emp_no not in svc_staff_sets[svc]:
                    svc_staff_sets[svc].add(emp_no)
                    g['staff_count'] += 1
                    g['total_work_hours'] += work_h
                    if eff > 0:
                        g['efficiencies'].append(eff)
                    if ach_rate > 0:
                        g['achievement_rates'].append(ach_rate)
                    if is_achieved:
                        g['achieved_count'] += 1


        # 计算每个服务标准分组的汇总指标
        for svc, g in svc_groups.items():
            g['avg_efficiency'] = round(sum(g['efficiencies']) / len(g['efficiencies']), 1) if g['efficiencies'] else 0
            g['avg_work_hours'] = round(g['total_work_hours'] / g['staff_count'], 1) if g['staff_count'] > 0 else 0
            g['achievement_rate'] = round(sum(g['achievement_rates']) / len(g['achievement_rates']), 1) if g['achievement_rates'] else 0
            g['upph'] = round(g['total_count'] / g['total_work_hours'], 1) if g['total_work_hours'] > 0 else 0

        # ===== 瑕疵图拍照完成量（flaw_total 已在前面获取并计入瑕疵图岗位人员）=====

        # ===== 拍照方式分组 =====
        method_groups = {
            '哥斯拉拍照': {'method': '哥斯拉拍照', 'total_count': 0, 'staff_count': 0, 'efficiencies': [], 'upph': 0, 'total_hours': 0},
            '拍照魔方': {'method': '拍照魔方', 'total_count': 0, 'staff_count': 0, 'efficiencies': [], 'upph': 0, 'total_hours': 0},
            '光合拍照': {'method': '光合拍照', 'total_count': 0, 'staff_count': 0, 'efficiencies': [], 'upph': 0, 'total_hours': 0},
            '瑕疵图拍照': {'method': '瑕疵图拍照', 'total_count': 0, 'staff_count': 0, 'efficiencies': [], 'upph': 0, 'total_hours': 0},
        }
        for emp_no, s in photo_staff.items():
            sources = s.get('data_sources', [])
            if 'godzilla' in sources and s.get('godzilla_count', 0) > 0:
                mg = method_groups['哥斯拉拍照']
                mg['total_count'] += s['godzilla_count']
                mg['staff_count'] += 1
                mg['total_hours'] += s['work_hours']
                if s['efficiency'] > 0:
                    mg['efficiencies'].append(s['efficiency'])
            if s.get('photo_cube_count', 0) > 0:
                mg = method_groups['拍照魔方']
                mg['total_count'] += s['photo_cube_count']
                mg['staff_count'] += 1
                mg['total_hours'] += s['work_hours']
                if s['efficiency'] > 0:
                    mg['efficiencies'].append(s['efficiency'])
            if 'guanghe' in sources:
                mg = method_groups['光合拍照']
                gh_cnt = s.get('guanghe_categories', {})
                gh_total = sum(v.get('count', 0) if isinstance(v, dict) else v for v in gh_cnt.values())
                mg['total_count'] += gh_total
                mg['staff_count'] += 1
                mg['total_hours'] += s['work_hours']
                if s['efficiency'] > 0:
                    mg['efficiencies'].append(s['efficiency'])

        # 瑕疵图拍照：体积指标，直接用 widget 7757 的完成量填充
        from config import FLAW_PHOTO_STANDARDS
        mg = method_groups['瑕疵图拍照']
        mg['standard'] = FLAW_PHOTO_STANDARDS  # 前端可展示达标标准
        if flaw_total > 0:
            mg['total_count'] = flaw_total
            # 瑕疵图人员：当日实际拍了瑕疵图(瑕疵图明细出现其姓名)的拍照组人员，
            # 不再限于 position_tag='瑕疵图' 一人（纠正此前只计欧紫月的错误）
            flaw_staff = [s for s in photo_staff.values() if s.get('flaw_count', 0) > 0]
            mg['staff_count'] = len(flaw_staff)
            mg['total_hours'] = round(sum(s['work_hours'] for s in flaw_staff), 1)

        for mk, mg in method_groups.items():
            mg['avg_efficiency'] = round(sum(mg['efficiencies']) / len(mg['efficiencies']), 1) if mg['efficiencies'] else 0
            mg['upph'] = round(mg['total_count'] / mg['total_hours'], 1) if mg['total_hours'] > 0 else 0

        # ===== 排序 + 标记 =====
        staff_list = sorted(
            photo_staff.values(),
            key=lambda x: x.get('achievement_rate', 0),
            reverse=True
        )
        for i, s in enumerate(staff_list):
            s['rank'] = i + 1
            s['is_top'] = (i == 0 and s.get('achievement_rate', 0) > 0)

        # ===== 整体统计 =====
        total_count = sum(s['total_count'] for s in staff_list)
        total_hours = sum(s['work_hours'] for s in staff_list)
        avg_eff = round(total_count / total_hours, 1) if total_hours > 0 else 0
        rates = [s['achievement_rate'] for s in staff_list if s['achievement_rate'] > 0]
        avg_rate = round(sum(rates) / len(rates), 1) if rates else 0
        achieved_count = sum(1 for s in staff_list if s['achievement_rate'] >= 100)

        # 单小时运能汇总（后端计算，避免 Jinja2 for 循环内 set 作用域 bug）
        from collections import defaultdict
        _hourly_col = defaultdict(int)
        _hourly_grand = 0
        for s in staff_list:
            hc = s.get('hourly_counts') or {}
            for h, v in hc.items():
                _hourly_col[str(h)] += int(v)
            _hourly_grand += sum(int(v) for v in hc.values())
        hourly_summary = {'by_hour': dict(_hourly_col), 'total': _hourly_grand}

        result = {
            'date': target_date,
            'data_source': 'multi',
            'total_count': total_count,
            'total_work_hours': round(total_hours, 1),
            'avg_efficiency': avg_eff,
            'achievement_rate': avg_rate,
            'staff_count': len(staff_list),
            'achieved_count': achieved_count,
            'staff_ranking': staff_list,
            'service_standard_groups': list(svc_groups.values()),
            'method_groups': list(method_groups.values()),
            'group_count': len(svc_groups),
            'session_expired': getattr(self.godzilla_client, 'session_expired', False),
            'hourly_summary': hourly_summary,
        }

        _cache.set(cache_key, result)
        return result

    def _calc_auto_target(self, s: Dict, data_source: str) -> float:
        """
        自动化(哥斯拉/魔镜)动态达标目标：返回各小时目标的平均值（综合达成的分母）。

        规则（用户 2026-07-19）：
        哥斯拉：
          - 单台设备(mc==1): 全职 58 / 兼职 52（平铺，无逐小时动态）
          - 双台及以上(mc>=2): 逐小时动态 —— 单小时处理量 >= 70 → 135，否则 → 58
        魔镜：
          - 两台设备(mc==2): 123
          - 三台设备(mc==3): 183.7
          - 其它设备数: 回退 183.7（魔镜默认，数据里目前只有 2/3 台）

        每个小时的 [小时, 处理量, 目标, 该小时达成] 写入 s['hourly_detail'] 供前端展示。
        """
        hours = s.get('hourly_counts', {}) or {}
        emp_type = s.get('employee_type', '') or ''
        # 哥斯拉单台平铺标准（按全职/兼职区分）
        godzilla_single = (EFFICIENCY_CONFIG['fulltime_standard'] if '全职' in emp_type
                           else EFFICIENCY_CONFIG['parttime_standard'])
        machine_count = int(s.get('machine_count', 1) or 1)

        def _target_for_hour(vol):
            if data_source == 'godzilla':
                if machine_count >= 2:
                    return 135.0 if vol >= 70 else 58.0
                return float(godzilla_single)
            if data_source == 'mirror':
                if machine_count == 2:
                    return 123.0
                if machine_count == 3:
                    return 183.7
                return 183.7  # 魔镜其它设备数默认值
            return float(godzilla_single)

        if not hours:
            # 回退：无小时数据时平铺
            return float(_target_for_hour(0))

        targets = []
        hourly_detail = []
        for h, vol in hours.items():
            t = _target_for_hour(vol)
            targets.append(t)
            hourly_detail.append({
                'hour': h,
                'volume': vol,
                'target': t,
                'achievement': round(vol / t * 100, 1) if t > 0 else 0,
            })
        avg_target = round(sum(targets) / len(targets), 1)
        s['hourly_detail'] = sorted(hourly_detail, key=lambda x: x['hour'])
        return avg_target

    def get_automation_efficiency(self, target_date: str = None) -> Dict:
        """
        获取自动化效率（哥斯拉+魔镜双数据源）
        - 哥斯拉系统的数据 → 哥斯拉组
        - 魔镜系统的数据 → 魔镜组
        - 人员按花名册position_tag筛选（只保留哥斯拉/魔镜标签的人）
        - 同一人在两边都有数据时，各算各的，分别在两个组里展示
        """
        if target_date is None:
            target_date = date.today().strftime('%Y-%m-%d')

        cache_key = self._get_cache_key('automation_efficiency', target_date)
        cached = _cache.get(cache_key)
        if cached is not None:
            return cached

        # 每次刷新重新加载花名册，确保人员管理改动实时生效
        self.reload_roster()

        # 同时拉取哥斯拉和魔镜数据
        godzilla_records = self.get_godzilla_records(target_date)
        mirror_records = self.get_mirror_records(target_date)

        # 分别聚合（_aggregate_by_staff返回{emp_no: staff_data}）
        godzilla_staff = self._aggregate_by_staff(godzilla_records, 'godzilla')
        mirror_staff = self._aggregate_by_staff(mirror_records, 'mirror')

        # 最终人员列表的key = emp_no + "_" + data_source，避免同一人两边数据互相覆盖
        auto_staff = {}

        # === 处理哥斯拉数据 ===
        for emp_no, s in godzilla_staff.items():
            info = get_staff_info(emp_no, self.roster)
            pos_tag = info.get('position_tag', '')
            # 只保留position_tag为哥斯拉/魔镜的人员
            if pos_tag not in ('哥斯拉', '魔镜'):
                continue
            # 补全信息
            s['name'] = info.get('name', s.get('name', emp_no))
            s['employee_type'] = info.get('employee_type', '')
            s['position_tag'] = pos_tag
            s['group_name'] = '哥斯拉'
            s['data_source'] = 'godzilla'
            # 计算达标率（哥斯拉：按小时动态目标，双台且单小时>80→135，否则58）
            s['target_efficiency'] = self._calc_auto_target(s, 'godzilla')
            if s['target_efficiency'] > 0 and s.get('efficiency', 0) > 0:
                s['achievement_rate'] = round(s['efficiency'] / s['target_efficiency'] * 100, 1)
            else:
                s['achievement_rate'] = 0
            key = emp_no + '_godzilla'
            auto_staff[key] = s

        # === 处理魔镜数据 ===
        for emp_no, s in mirror_staff.items():
            info = get_staff_info(emp_no, self.roster)
            pos_tag = info.get('position_tag', '')
            # 只保留position_tag为哥斯拉/魔镜的人员
            if pos_tag not in ('哥斯拉', '魔镜'):
                continue
            # 补全信息
            s['name'] = info.get('name', s.get('name', emp_no))
            s['employee_type'] = info.get('employee_type', '')
            s['position_tag'] = pos_tag
            s['group_name'] = '魔镜'
            s['data_source'] = 'mirror'
            # 计算达标率（魔镜：两台设备→123，否则沿用183.7，按小时动态目标均值）
            s['target_efficiency'] = self._calc_auto_target(s, 'mirror')
            if s['target_efficiency'] > 0 and s.get('efficiency', 0) > 0:
                s['achievement_rate'] = round(s['efficiency'] / s['target_efficiency'] * 100, 1)
            else:
                s['achievement_rate'] = 0
            key = emp_no + '_mirror'
            auto_staff[key] = s

        # === 组装结果 ===
        staff_list = sorted(
            auto_staff.values(),
            key=lambda x: x.get('achievement_rate', 0),
            reverse=True
        )
        for i, s in enumerate(staff_list):
            s['rank'] = i + 1
            s['is_top'] = (i == 0 and s.get('achievement_rate', 0) > 0)

        # 汇总
        total_count = sum(s.get('total_count', 0) for s in staff_list)
        total_hours = sum(s.get('work_hours', 0) for s in staff_list)
        avg_eff = round(total_count / total_hours, 1) if total_hours > 0 else 0
        rates = [s['achievement_rate'] for s in staff_list if s.get('achievement_rate', 0) > 0]
        avg_rate = round(sum(rates) / len(rates), 1) if rates else 0

        # 分组统计
        groups = []
        for group_name in ['哥斯拉', '魔镜']:
            group_staff = [s for s in staff_list if s.get('group_name') == group_name]
            g_count = sum(s.get('total_count', 0) for s in group_staff)
            g_hours = sum(s.get('work_hours', 0) for s in group_staff)
            g_avg_eff = round(g_count / g_hours, 1) if g_hours > 0 else 0
            g_rates = [s['achievement_rate'] for s in group_staff if s.get('achievement_rate', 0) > 0]
            g_avg_rate = round(sum(g_rates) / len(g_rates), 1) if g_rates else 0
            g_achieved = sum(1 for s in group_staff if s.get('achievement_rate', 0) >= 100)
            groups.append({
                'group_name': group_name,
                'total_count': g_count,
                'staff_count': len(group_staff),
                'achieved_count': g_achieved,
                'achievement_rate': g_avg_rate,
                'avg_efficiency': g_avg_eff,
                'total_work_hours': round(g_hours, 1),
            })

        session_expired = self.godzilla_client.session_expired or self.mirror_client.session_expired

        result = {
            'date': target_date,
            'data_source': 'godzilla+mirror',
            'total_count': total_count,
            'staff_count': len(staff_list),
            'avg_efficiency': avg_eff,
            'achievement_rate': avg_rate,
            'total_work_hours': round(total_hours, 1),
            'staff_list': staff_list,
            'staff_ranking': staff_list,
            'groups': groups,
            'session_expired': session_expired,
        }

        _cache.set(cache_key, result)
        return result




    def get_xray_efficiency(self, date_str=None):
        """
        获取X光设备效率（按设备维度，不是人员维度）
        返回结构和automation一致，方便前端复用
        """
        from datetime import datetime
        import config as _cfg
        
        if not date_str:
            date_str = datetime.now().strftime('%Y-%m-%d')
        
        fmt = '%Y-%m-%d %H:%M:%S'
        standard = _cfg.XRAY_STANDARD_UPH
        machines = _cfg.XRAY_CHENGDU_MACHINES
        
        machine_stats = []
        total_count = 0
        total_work_hours = 0.0
        achieved_count = 0
        
        for mid in machines:
            try:
                records = self._get_xray_client().fetch_machine_records(mid, date_str)
            except Exception as e:
                print(f"[XRay] 拉取{mid}失败: {e}")
                records = []
            
            if not records:
                machine_stats.append({
                    'machine_id': mid,
                    'name': mid,
                    'total_count': 0,
                    'work_hours': 0,
                    'upph': 0,
                    'achievement_rate': 0,
                    'achieved': False,
                    'hourly_counts': {},
                })
                continue
            
            count = len(records)
            
            # 计算首末时间
            start_times = [r.get('startDt', '') for r in records if r.get('startDt')]
            end_times = [r.get('endDt', '') for r in records if r.get('endDt')]
            
            first_start = min(start_times) if start_times else ''
            last_end = max(end_times) if end_times else ''
            
            work_hours = 0.0
            if first_start and last_end:
                try:
                    t1 = datetime.strptime(first_start, fmt)
                    t2 = datetime.strptime(last_end, fmt)
                    work_hours = (t2 - t1).total_seconds() / 3600.0
                    # X光除外：不参与午休剔除（用户 2026-07-19 明确规则），
                    # 即便结束时间 >=13:30 也不减 1 小时。
                except:
                    pass
            
            upph = count / work_hours if work_hours > 0 else 0
            achievement_rate = round(upph / standard * 100, 1) if standard > 0 else 0
            achieved = upph >= standard
            
            if achieved:
                achieved_count += 1
            
            total_count += count
            total_work_hours += work_hours
            
            # 时段统计
            hourly_counts = {}
            for r in records:
                sd = r.get('startDt', '')
                if sd and len(sd) >= 13:
                    hour = sd[11:13]
                    hourly_counts[hour] = hourly_counts.get(hour, 0) + 1
            
            machine_stats.append({
                'machine_id': mid,
                'name': mid,
                'total_count': count,
                'work_hours': round(work_hours, 2),
                'upph': round(upph, 1),
                'achievement_rate': achievement_rate,
                'achieved': achieved,
                'hourly_counts': hourly_counts,
            })
        
        # 按UPPH排序
        machine_stats.sort(key=lambda x: x['upph'], reverse=True)
        
        avg_upph = total_count / total_work_hours if total_work_hours > 0 else 0
        total_machines = len([m for m in machine_stats if m['total_count'] > 0])
        achievement_rate = round(achieved_count / total_machines * 100, 1) if total_machines > 0 else 0
        
        # groups结构（和automation兼容）
        groups = [{
            'group_name': 'X-RAY',
            'total_count': total_count,
            'staff_count': total_machines,  # 复用字段，实际是设备数
            'achieved_count': achieved_count,
            'achievement_rate': achievement_rate,
            'avg_efficiency': round(avg_upph, 1),
            'total_work_hours': round(total_work_hours, 2),
        }]
        
        return {
            'date': date_str,
            'data_source': 'xray',
            'total_count': total_count,
            'staff_count': total_machines,  # 设备数
            'avg_efficiency': round(avg_upph, 1),
            'achievement_rate': achievement_rate,
            'total_work_hours': round(total_work_hours, 2),
            'achieved_count': achieved_count,
            'staff_list': machine_stats,  # 设备列表
            'staff_ranking': machine_stats,
            'groups': groups,
            'machines': machine_stats,
            'session_expired': False,
        }

    def _recalc_efficiency(self, s, pos_tag):
        """重新计算UPPH和达标率（按position_tag对应的标准）"""
        # 工作时长已经在_aggregate_by_staff里算好了，直接用
        work_hours = s.get('work_hours', 0)
        total_count = s.get('total_count', 0)
        if work_hours > 0:
            s['efficiency'] = round(total_count / work_hours, 1)
        else:
            s['efficiency'] = 0
        # 达标标准
        emp_type = s.get('employee_type', '')
        if pos_tag == '魔镜':
            target = EFFICIENCY_CONFIG['mirror_fulltime_standard']  # 魔镜统一183.7
        elif '全职' in emp_type:
            target = EFFICIENCY_CONFIG['fulltime_standard']
        else:
            target = EFFICIENCY_CONFIG['parttime_standard']
        s['target_efficiency'] = target
        if target > 0 and s['efficiency'] > 0:
            s['achievement_rate'] = round(s['efficiency'] / target * 100, 1)
        else:
            s['achievement_rate'] = 0


    def get_app_efficiency(self, target_date=None):
        """获取APP效率数据 - 从金山文档v3 API下载xlsx解析"""
        import traceback
        from datetime import datetime, timedelta
        
        result = {
            'date': target_date or datetime.now().strftime('%Y-%m-%d'),
            'total_count': 0,
            'staff_count': 0,
            'avg_efficiency': 0,
            'achievement_rate': 0,
            'total_work_hours': 0,
            'staff_ranking': [],
            'source': 'kdocs',
            'status': 'error',
            'message': '',
        }
        
        try:
            if target_date is None:
                target_date = datetime.now().strftime('%Y-%m-%d')
            result['date'] = target_date
            
            # 初始化金山文档客户端
            if not hasattr(self, '_kdocs') or self._kdocs is None:
                from services.kdocs_client import KdocsClient
                from config import KDOCS_CONFIG
                self._kdocs = KdocsClient(
                    token=KDOCS_CONFIG.get('token', ''),
                    folder_share_url=KDOCS_CONFIG.get('folder_share_url', '')
                )
                cookie = KDOCS_CONFIG.get('cookie', '')
                if cookie:
                    if hasattr(self._kdocs, 'set_cookie'):
                        self._kdocs.set_cookie(cookie)
                    else:
                        self._kdocs.session.headers['Cookie'] = cookie
                teamid = KDOCS_CONFIG.get('teamid', '')
                if teamid and hasattr(self._kdocs, 'set_teamid'):
                    self._kdocs.set_teamid(teamid)
            
            # 列文件
            if hasattr(self._kdocs, 'list_team_files'):
                files = self._kdocs.list_team_files(limit=200)
                logger.info(f'[APP效率] 金山文档文件列表: {len(files)}个')
            else:
                files = []
                logger.warning('[APP效率] kdocs客户端没有list_team_files方法')
            
            if not files:
                result['status'] = 'no_files'
                result['message'] = '未获取到文件列表'
                return result
            
            xlsx_files = [f for f in files if (f.get('fname') or f.get('name','')).endswith(('.xlsx','.xls','.et'))]
            logger.info(f'[APP效率] xlsx文件数: {len(xlsx_files)}个')
            
            import openpyxl
            staff_list = []
            
            for _idx, f in enumerate(xlsx_files):
                fname = f.get('fname') or f.get('name', '')
                file_id = f.get('id')
                if not file_id:
                    continue
                try:
                    logger.info(f'[APP效率] 处理 {_idx+1}/{len(xlsx_files)}: {fname[:20]}')
                    import signal
                    dl_path = None
                    try:
                        dl_path = self._kdocs.download_xlsx(file_id)
                    except Exception as dl_e:
                        logger.warning(f'[APP效率] 下载失败 {fname}: {dl_e}')
                        continue
                    if not dl_path:
                        logger.warning(f'[APP效率] 下载失败: {fname[:20]}')
                        continue
                    person_data = self._parse_app_xlsx(dl_path, target_date)
                    if not person_data.get('staff_name') or person_data.get('staff_name', '').startswith('kdocs') or person_data.get('staff_name') == 'kdocs':
                        name = fname.replace('.xlsx','').replace('.xls','').replace('.et','')
                        if '——' in name:
                            name = name.split('——')[0]
                        elif '-' in name:
                            name = name.split('-')[0]
                        person_data['staff_name'] = name.strip()
                    staff_list.append(person_data)
                    try:
                        import os
                        os.remove(dl_path)
                    except:
                        pass
                except Exception as e:
                    print(f"[app_eff] 处理 {fname} 失败: {e}")
                    continue
            
            logger.info(f'[APP效率] 成功解析 {len(staff_list)} 人数据')
            today_count = sum(s.get('today_count', 0) for s in staff_list)
            month_count = sum(s.get('month_count', 0) for s in staff_list)
            # 当日真实处理量：使用按日期筛选的 today_count（已为『今日』维度）
            
            staff_list.sort(key=lambda x: x.get('today_count', 0), reverse=True)
            staff_ranking = []
            for s in staff_list:
                staff_ranking.append({
                    'name': s.get('staff_name', ''),
                    'today': s.get('today_count', 0),
                    'month': s.get('month_count', 0),
                    'total': s.get('total_records', 0),
                    'upph': s.get('upph', 0),
                    'hourly_counts': s.get('hourly_counts', {}),
                    'achievement_rate': 0,
                    # ---- 保留完整解析字段（扫码时间/设备类型/工时）----
                    'android_count': s.get('android_count', 0),
                    'ios_count': s.get('ios_count', 0),
                    'android_first': self._dt_str(s.get('android_first')),
                    'android_last': self._dt_str(s.get('android_last')),
                    'ios_first': self._dt_str(s.get('ios_first')),
                    'ios_last': self._dt_str(s.get('ios_last')),
                    'overall_first': self._dt_str(s.get('overall_first')),
                    'overall_last': self._dt_str(s.get('overall_last')),
                    'android_hours': s.get('android_hours', 0),
                    'ios_hours': s.get('ios_hours', 0),
                    'work_hours': s.get('work_hours', 0),
                    'work_type': s.get('work_type', '安卓'),
                    # ---- 当日维度：今日首末扫码时间 / 当日真实处理量 ----
                    'today_real': s.get('today_count', 0),
                    'today_overall_first': self._dt_str(s.get('today_overall_first')),
                    'today_overall_last': self._dt_str(s.get('today_overall_last')),
                    # ---- 当日装跑类型判断依据：按今日是否有各表扫码记录 ----
                    'today_android_first': self._dt_str(s.get('today_android_first')),
                    'today_android_last': self._dt_str(s.get('today_android_last')),
                    'today_ios_first': self._dt_str(s.get('today_ios_first')),
                    'today_ios_last': self._dt_str(s.get('today_ios_last')),
                })
            
            result['total_count'] = today_count
            result['staff_count'] = len(staff_list)
            result['avg_efficiency'] = round(today_count / len(staff_list), 1) if staff_list else 0
            result['staff_ranking'] = staff_ranking
            result['today_count'] = today_count
            result['month_count'] = month_count
            result['status'] = 'ok'
            return result
            
        except Exception as e:
            traceback.print_exc()
            result['status'] = 'error'
            result['message'] = str(e)
            return result

    def _parse_app_xlsx(self, xlsx_path, target_date=None):
        """解析单个APP登记xlsx文件 - 表头定位法
        先找表头行（含"扫码时间"/"日期"文字），根据表头定位列，再读数据。
        target_date: 用于从带日期的扫码记录中筛选『当日』维度（首末扫码时间/计数/逐小时）。
        """
        import os as _os
        import openpyxl
        from datetime import datetime, date, time
        
        result = {
            'staff_name': '',
            'today_count': 0,
            'month_count': 0,
            'yesterday_count': 0,
            'total_records': 0,
            'upph': 0,
            'android_count': 0,
            'ios_count': 0,
            'total_count': 0,
            'android_first': None,
            'android_last': None,
            'ios_first': None,
            'ios_last': None,
            'overall_first': None,
            'overall_last': None,
            'android_hours': 0,
            'ios_hours': 0,
            'work_hours': 0,
            'hourly_counts': {},
            'work_type': '安卓',
            # ---- 当日维度（按 target_date 筛选的『今日』真实数据）----
            'today_real_count': 0,
            'today_overall_first': None,
            'today_overall_last': None,
            'today_android_first': None,
            'today_android_last': None,
            'today_ios_first': None,
            'today_ios_last': None,
            'today_hourly': {},
        }
        
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            
            # 从文件名提取员工姓名（排除kdocs临时文件）
            fname_base = _os.path.basename(xlsx_path)
            name_extracted = False
            for sep in ['——', '--', '-', '_']:
                if sep in fname_base:
                    candidate = fname_base.split(sep)[0].strip()
                    if not candidate.startswith('kdocs_') and len(candidate) > 1 and candidate != 'kdocs':
                        result['staff_name'] = candidate
                        name_extracted = True
                        break
            
            # 从A1单元格读姓名兜底
            if not name_extracted:
                try:
                    ws_first = wb[wb.sheetnames[0]]
                    a1_val = ws_first.cell(row=1, column=1).value
                    if a1_val and isinstance(a1_val, str) and len(a1_val) < 20:
                        result['staff_name'] = a1_val.strip()
                except:
                    pass
            
            # 分类sheet
            android_sheets = []
            ios_sheets = []
            for s in wb.sheetnames:
                s_low = s.lower()
                if 'app' in s_low and ('登记' in s or '效率' in s or '统计' in s or '装跑' in s):
                    android_sheets.append(s)
                elif '苹果' in s or '测跑' in s or 'ios' in s_low:
                    ios_sheets.append(s)
            
            if not android_sheets and not ios_sheets and wb.sheetnames:
                android_sheets = [wb.sheetnames[0]]
            
            def _parse_sheet(ws, target_date=None):
                """解析单个sheet，返回(count, first_dt, last_dt, hourly,
                today_count, today_first, today_last, today_hourly)"""
                # 第1步：找表头行（包含"扫码时间"或"日期"字样）
                header_row = None
                datetime_col = None
                date_col = None
                
                for scan_row in range(1, min(20, ws.max_row + 1)):
                    found_any = False
                    for scan_col in range(1, min(15, ws.max_column + 1)):
                        v = ws.cell(row=scan_row, column=scan_col).value
                        if isinstance(v, str):
                            v_str = v.strip()
                            if '扫码时间' in v_str:
                                datetime_col = scan_col
                                found_any = True
                            elif v_str == '日期':
                                date_col = scan_col
                                found_any = True
                    if found_any:
                        header_row = scan_row
                        break
                
                if header_row is None:
                    header_row = 4
                    if datetime_col is None:
                        datetime_col = 3
                    if date_col is None:
                        date_col = 4
                
                start_row = header_row + 1
                count = 0
                first_dt = None
                last_dt = None
                hourly_counts = {}

                # 今日维度（按 target_date 的日期筛选扫码记录）
                _today_date = None
                if target_date:
                    if isinstance(target_date, str):
                        try:
                            _today_date = datetime.strptime(target_date[:10], '%Y-%m-%d').date()
                        except Exception:
                            _today_date = None
                    elif isinstance(target_date, date):
                        _today_date = target_date
                today_count = 0
                today_first = None
                today_last = None
                today_hourly = {}
                
                for row_idx in range(start_row, ws.max_row + 1):
                    try:
                        row_dt = None
                        if datetime_col:
                            v = ws.cell(row=row_idx, column=datetime_col).value
                            if isinstance(v, datetime):
                                row_dt = v
                            elif isinstance(v, date):
                                row_dt = datetime.combine(v, time(0, 0, 0))
                            elif isinstance(v, str) and len(v) >= 10:
                                for fmt in ['%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S',
                                            '%Y-%m-%d %H:%M', '%Y/%m/%d %H:%M']:
                                    try:
                                        s = v.strip()
                                        row_dt = datetime.strptime(s[:19] if len(s) >= 19 else s, fmt)
                                        break
                                    except:
                                        pass
                        if not row_dt and date_col:
                            d = ws.cell(row=row_idx, column=date_col).value
                            if isinstance(d, datetime):
                                row_dt = d
                            elif isinstance(d, date):
                                row_dt = datetime.combine(d, time(0, 0, 0))
                        
                        if not row_dt:
                            continue
                        count += 1
                        h = row_dt.hour
                        hourly_counts[h] = hourly_counts.get(h, 0) + 1
                        if first_dt is None or row_dt < first_dt:
                            first_dt = row_dt
                        if last_dt is None or row_dt > last_dt:
                            last_dt = row_dt
                        # 今日维度：仅统计扫码日期 == target_date 的记录
                        if _today_date and row_dt.date() == _today_date:
                            today_count += 1
                            today_hourly[h] = today_hourly.get(h, 0) + 1
                            if today_first is None or row_dt < today_first:
                                today_first = row_dt
                            if today_last is None or row_dt > today_last:
                                today_last = row_dt
                    except:
                        continue
                
                return count, first_dt, last_dt, hourly_counts, today_count, today_first, today_last, today_hourly
            
            # 解析安卓sheet
            for sheet_name in android_sheets:
                ws = wb[sheet_name]
                cnt, first, last, hourly, tcnt, tf, tl, th = _parse_sheet(ws, target_date)
                result['android_count'] += cnt
                for h, v in hourly.items():
                    result['hourly_counts'][h] = result['hourly_counts'].get(h, 0) + v
                if first:
                    if result['android_first'] is None or first < result['android_first']:
                        result['android_first'] = first
                if last:
                    if result['android_last'] is None or last > result['android_last']:
                        result['android_last'] = last
                # 当日维度累加
                result['today_real_count'] += tcnt
                for h, v in th.items():
                    result['today_hourly'][h] = result['today_hourly'].get(h, 0) + v
                if tf:
                    if result['today_android_first'] is None or tf < result['today_android_first']:
                        result['today_android_first'] = tf
                if tl:
                    if result['today_android_last'] is None or tl > result['today_android_last']:
                        result['today_android_last'] = tl
            
            # 解析苹果sheet
            for sheet_name in ios_sheets:
                ws = wb[sheet_name]
                cnt, first, last, hourly, tcnt, tf, tl, th = _parse_sheet(ws, target_date)
                result['ios_count'] += cnt
                for h, v in hourly.items():
                    result['hourly_counts'][h] = result['hourly_counts'].get(h, 0) + v
                if first:
                    if result['ios_first'] is None or first < result['ios_first']:
                        result['ios_first'] = first
                if last:
                    if result['ios_last'] is None or last > result['ios_last']:
                        result['ios_last'] = last
                # 当日维度累加
                result['today_real_count'] += tcnt
                for h, v in th.items():
                    result['today_hourly'][h] = result['today_hourly'].get(h, 0) + v
                if tf:
                    if result['today_ios_first'] is None or tf < result['today_ios_first']:
                        result['today_ios_first'] = tf
                if tl:
                    if result['today_ios_last'] is None or tl > result['today_ios_last']:
                        result['today_ios_last'] = tl
            
            wb.close()
            
            result['total_count'] = result['android_count'] + result['ios_count']
            result['today_count'] = result['today_real_count']
            # 当日逐小时产量覆盖为『今日』维度（供页面时段表使用）
            # 无论今日是否有扫码记录，都用 today_hourly（无数据时为空字典{}，不回退整月）
            result['hourly_counts'] = result.get('today_hourly') or {}
            
            if result['ios_count'] > 0 and result['android_count'] == 0:
                result['work_type'] = '苹果'
            elif result['ios_count'] > 0 and result['android_count'] > 0:
                result['work_type'] = '安卓+苹果'
            
            all_firsts = [t for t in [result['android_first'], result['ios_first']] if t is not None]
            all_lasts = [t for t in [result['android_last'], result['ios_last']] if t is not None]
            if all_firsts:
                result['overall_first'] = min(all_firsts)
            if all_lasts:
                result['overall_last'] = max(all_lasts)

            # 当日首末扫码时间（仅 target_date 当天的记录）
            t_firsts = [t for t in [result['today_android_first'], result['today_ios_first']] if t is not None]
            t_lasts = [t for t in [result['today_android_last'], result['today_ios_last']] if t is not None]
            if t_firsts:
                result['today_overall_first'] = min(t_firsts)
            if t_lasts:
                result['today_overall_last'] = max(t_lasts)
            
            result['android_hours'] = self._calc_app_work_hours(
                result['android_first'], result['android_last'])
            result['ios_hours'] = self._calc_app_work_hours(
                result['ios_first'], result['ios_last'])
            result['work_hours'] = result['android_hours'] + result['ios_hours']
            
            if result['work_hours'] > 0 and result['total_count'] > 0:
                result['upph'] = round(result['total_count'] / result['work_hours'], 2)
        
        except Exception as e:
            print(f"[parse_app_xlsx] 解析失败: {e}")
            import traceback
            traceback.print_exc()
        
        return result

    def _calc_app_work_hours(self, first_time, last_time):
        """计算APP岗位工作时长（首台到末台，12:00-13:00午休扣1小时）"""
        if first_time is None or last_time is None:
            return 0
        from datetime import datetime, time
        if isinstance(first_time, str):
            first_time = datetime.strptime(first_time, '%Y-%m-%d %H:%M:%S')
        if isinstance(last_time, str):
            last_time = datetime.strptime(last_time, '%Y-%m-%d %H:%M:%S')
        total_seconds = (last_time - first_time).total_seconds()
        total_hours = total_seconds / 3600
        # 末台在13:00之后且首台在12:00之前，扣1小时午休
        if last_time.time() >= time(13, 0) and first_time.time() <= time(12, 0):
            total_hours -= 1
        return round(max(total_hours, 0), 2)

    @staticmethod
    def _dt_str(dt):
        """datetime对象转字符串（用于JSON序列化），None/非dt返回None"""
        if dt is None:
            return None
        if isinstance(dt, str):
            return dt
        try:
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        except Exception:
            return None


    def get_group_efficiency(self, target_date: str = None) -> Dict:
        """
        获取所有组别的效率汇总
        """
        if target_date is None:
            target_date = date.today().strftime('%Y-%m-%d')

        cache_key = self._get_cache_key('group_efficiency', target_date)
        cached = _cache.get(cache_key)
        if cached is not None:
            return cached

        photo_data = self.get_photo_efficiency(target_date)
        auto_data = self.get_automation_efficiency(target_date)
        app_data = self.get_app_efficiency(target_date)

        # 合并所有组别
        all_groups = {}

        for g in photo_data.get('groups', []):
            name = g['group_name']
            if name not in all_groups:
                all_groups[name] = {
                    'group_name': name,
                    'photo': {'total_count': 0, 'staff_count': 0, 'achievement_rate': 0, 'avg_efficiency': 0},
                    'automation': {'total_count': 0, 'staff_count': 0, 'achievement_rate': 0, 'avg_efficiency': 0},
                    'app': {'total_count': 0, 'staff_count': 0, 'achievement_rate': 0, 'avg_efficiency': 0},
                }
            all_groups[name]['photo'] = {
                'total_count': g['total_count'],
                'staff_count': g['staff_count'],
                'achievement_rate': g['achievement_rate'],
                'avg_efficiency': g['avg_efficiency'],
            }

        for g in auto_data.get('groups', []):
            name = g['group_name']
            if name not in all_groups:
                all_groups[name] = {
                    'group_name': name,
                    'photo': {'total_count': 0, 'staff_count': 0, 'achievement_rate': 0, 'avg_efficiency': 0},
                    'automation': {'total_count': 0, 'staff_count': 0, 'achievement_rate': 0, 'avg_efficiency': 0},
                    'app': {'total_count': 0, 'staff_count': 0, 'achievement_rate': 0, 'avg_efficiency': 0},
                }
            all_groups[name]['automation'] = {
                'total_count': g['total_count'],
                'staff_count': g['staff_count'],
                'achievement_rate': g['achievement_rate'],
                'avg_efficiency': g['avg_efficiency'],
            }

        result = {
            'date': target_date,
            'groups': list(all_groups.values()),
            'total_staff': sum(g.get('photo', {}).get('staff_count', 0) + g.get('automation', {}).get('staff_count', 0) + g.get('app', {}).get('staff_count', 0) for g in all_groups.values()),
            'total_output': sum(g.get('photo', {}).get('total_count', 0) + g.get('automation', {}).get('total_count', 0) + g.get('app', {}).get('total_count', 0) for g in all_groups.values()),
            'overall_achievement': round(sum(g.get('photo', {}).get('achievement_rate', 0) + g.get('automation', {}).get('achievement_rate', 0) for g in all_groups.values()) / max(len(all_groups)*2, 1), 1),
            'avg_efficiency': 0,
            'all_staff': [],
            'photo_summary': {
                'total_count': photo_data['total_count'],
                'staff_count': photo_data['staff_count'],
                'avg_efficiency': photo_data['avg_efficiency'],
                'achievement_rate': photo_data['achievement_rate'],
            },
            'automation_summary': {
                'total_count': auto_data['total_count'],
                'staff_count': auto_data['staff_count'],
                'avg_efficiency': auto_data['avg_efficiency'],
                'achievement_rate': auto_data['achievement_rate'],
            },
            'app_summary': {
                'total_count': app_data.get('total_count', 0),
                'staff_count': app_data.get('staff_count', 0),
                'avg_efficiency': app_data.get('avg_efficiency', 0),
                'achievement_rate': app_data.get('achievement_rate', 0),
            },
        }

        _cache.set(cache_key, result)
        return result

    def get_home_kpi(self) -> Dict:
        """
        首页KPI汇总
        包含今日、近7天、近30天的关键指标
        """
        cache_key = 'home_kpi'
        cached = _cache.get(cache_key)
        if cached is not None:
            return cached

        today = date.today()
        today_str = today.strftime('%Y-%m-%d')

        # 今日数据
        photo_today = self.get_photo_efficiency(today_str)
        auto_today = self.get_automation_efficiency(today_str)

        # 近7天
        seven_days_ago = (today - timedelta(days=6)).strftime('%Y-%m-%d')
        photo_7d = self._get_period_photo_efficiency(seven_days_ago, today_str)
        auto_7d = self._get_period_automation_efficiency(seven_days_ago, today_str)

        # 近30天
        thirty_days_ago = (today - timedelta(days=29)).strftime('%Y-%m-%d')
        photo_30d = self._get_period_photo_efficiency(thirty_days_ago, today_str)
        auto_30d = self._get_period_automation_efficiency(thirty_days_ago, today_str)

        result = {
            'today': {
                'date': today_str,
                'photo': {
                    'total_count': photo_today['total_count'],
                    'staff_count': photo_today['staff_count'],
                    'avg_efficiency': photo_today['avg_efficiency'],
                    'achievement_rate': photo_today['achievement_rate'],
                },
                'automation': {
                    'total_count': auto_today['total_count'],
                    'staff_count': auto_today['staff_count'],
                    'avg_efficiency': auto_today['avg_efficiency'],
                    'achievement_rate': auto_today['achievement_rate'],
                },
            },
            'last_7_days': {
                'photo': photo_7d,
                'automation': auto_7d,
            },
            'last_30_days': {
                'photo': photo_30d,
                'automation': auto_30d,
            },
            'session_expired': {
                'godzilla': self.godzilla_client.session_expired,
                'mirror': self.mirror_client.session_expired,
            },
        }

        _cache.set(cache_key, result)
        return result

    def _get_period_photo_efficiency(self, start_date: str, end_date: str) -> Dict:
        """获取一段时间内的拍照效率汇总（按天计算后汇总）"""
        cache_key = self._get_cache_key('period_photo', start_date, end_date)
        cached = _cache.get(cache_key)
        if cached is not None:
            return cached

        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()

        # 按天分别计算各员工的工作时长和产量，再汇总
        daily_emp_stats = {}  # emp_no -> {total_count, work_hours, target_efficiency}
        current = start
        while current <= end:
            date_str = current.strftime('%Y-%m-%d')
            records = self.get_godzilla_records(date_str)
            # 拍照周期汇总也只保留拍照类目
            from config import GODZILLA_PHOTO_CATEGORIES
            records = [
                r for r in records
                if str(r.get('service_standard', '') or r.get('serviceStandard', '')).strip()
                in GODZILLA_PHOTO_CATEGORIES
            ]
            day_staff = self._aggregate_by_staff(records, 'godzilla')
            for emp_no, s in day_staff.items():
                if emp_no not in daily_emp_stats:
                    daily_emp_stats[emp_no] = {
                        'total_count': 0,
                        'work_hours': 0,
                        'target_efficiency': s.get('target_efficiency', 0),
                        'achievement_days': 0,
                        'total_days': 0,
                    }
                daily_emp_stats[emp_no]['total_count'] += s['total_count']
                daily_emp_stats[emp_no]['work_hours'] += s['work_hours']
                if s['work_hours'] > 0:
                    daily_emp_stats[emp_no]['total_days'] += 1
                    if s['achievement_rate'] >= 100:
                        daily_emp_stats[emp_no]['achievement_days'] += 1
            current += timedelta(days=1)

        # 计算总体效率
        total_count = sum(s['total_count'] for s in daily_emp_stats.values())
        total_hours = sum(s['work_hours'] for s in daily_emp_stats.values())
        avg_eff = round(total_count / total_hours, 1) if total_hours > 0 else 0

        # 计算平均达标率（按员工平均）
        rates = []
        for emp_no, s in daily_emp_stats.items():
            if s['work_hours'] > 0 and s['target_efficiency'] > 0:
                emp_eff = s['total_count'] / s['work_hours']
                rates.append(round(emp_eff / s['target_efficiency'] * 100, 1))
        avg_rate = round(sum(rates) / len(rates), 1) if rates else 0

        result = {
            'start_date': start_date,
            'end_date': end_date,
            'total_count': total_count,
            'staff_count': len(daily_emp_stats),
            'avg_efficiency': avg_eff,
            'achievement_rate': avg_rate,
            'total_work_hours': round(total_hours, 1),
        }

        _cache.set(cache_key, result)
        return result

    def _get_period_automation_efficiency(self, start_date: str, end_date: str) -> Dict:
        """获取一段时间内的自动化效率汇总（按天计算后汇总）"""
        cache_key = self._get_cache_key('period_auto', start_date, end_date)
        cached = _cache.get(cache_key)
        if cached is not None:
            return cached

        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()

        # 按天分别计算各员工的工作时长和产量，再汇总
        daily_emp_stats = {}
        current = start
        while current <= end:
            date_str = current.strftime('%Y-%m-%d')
            records = self.get_mirror_records(date_str)
            day_staff = self._aggregate_by_staff(records, 'mirror')
            for emp_no, s in day_staff.items():
                if emp_no not in daily_emp_stats:
                    daily_emp_stats[emp_no] = {
                        'total_count': 0,
                        'work_hours': 0,
                        'target_efficiency': s.get('target_efficiency', 0),
                        'achievement_days': 0,
                        'total_days': 0,
                    }
                daily_emp_stats[emp_no]['total_count'] += s['total_count']
                daily_emp_stats[emp_no]['work_hours'] += s['work_hours']
                if s['work_hours'] > 0:
                    daily_emp_stats[emp_no]['total_days'] += 1
                    if s['achievement_rate'] >= 100:
                        daily_emp_stats[emp_no]['achievement_days'] += 1
            current += timedelta(days=1)

        total_count = sum(s['total_count'] for s in daily_emp_stats.values())
        total_hours = sum(s['work_hours'] for s in daily_emp_stats.values())
        avg_eff = round(total_count / total_hours, 1) if total_hours > 0 else 0

        rates = []
        for emp_no, s in daily_emp_stats.items():
            if s['work_hours'] > 0 and s['target_efficiency'] > 0:
                emp_eff = s['total_count'] / s['work_hours']
                rates.append(round(emp_eff / s['target_efficiency'] * 100, 1))
        avg_rate = round(sum(rates) / len(rates), 1) if rates else 0

        result = {
            'start_date': start_date,
            'end_date': end_date,
            'total_count': total_count,
            'staff_count': len(daily_emp_stats),
            'avg_efficiency': avg_eff,
            'achievement_rate': avg_rate,
            'total_work_hours': round(total_hours, 1),
        }

        _cache.set(cache_key, result)
        return result


# 全局服务实例
_efficiency_service: Optional[EfficiencyService] = None


def get_efficiency_service() -> EfficiencyService:
    """获取全局效率服务实例"""
    global _efficiency_service
    if _efficiency_service is None:
        _efficiency_service = EfficiencyService()
    return _efficiency_service
