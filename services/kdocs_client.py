# -*- coding: utf-8 -*-
"""
金山文档数据接入客户端 - 已修复v3方法 + 超时重定向保护
支持：v3 API 下载xlsx + 解析
"""
import os
import re
import json
import time
import urllib3
import requests
from datetime import datetime, timedelta

urllib3.disable_warnings()


class KdocsClient:
    """金山文档数据接入客户端"""

    def __init__(self, token='', folder_share_url=''):
        self.token = token
        self.folder_share_url = folder_share_url
        self.teamid = None
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'application/json, text/plain, */*',
        })

    # ============================================================
    # v3 API 方法（推荐，直接下载xlsx解析）
    # ============================================================
    def set_teamid(self, teamid):
        """设置团队/文件夹ID"""
        self.teamid = str(teamid)

    def set_cookie(self, wps_sid):
        """设置wps_sid cookie"""
        if not wps_sid:
            return
        # 清除header方式的cookie（避免冲突）
        if 'Cookie' in self.session.headers:
            del self.session.headers['Cookie']
        # 使用Session的cookies管理器，设置到多个相关域名
        for domain in ['.kdocs.cn', '.drive.kdocs.cn', 'drive.kdocs.cn']:
            self.session.cookies.set('wps_sid', wps_sid, domain=domain)
        print(f"[kdocs] set_cookie done: {wps_sid[:20]}...")

    def list_team_files(self, parentid='0', limit=200):
        """列出团队文件夹下的所有文件 - 多路径尝试（禁止重定向+JSON校验）"""
        if not self.teamid:
            return []
        # 尝试多种API路径
        api_paths = [
            # v3 groups API
            (f'https://drive.kdocs.cn/api/v3/groups/{self.teamid}/files',
             {'parentid': parentid, 'start': 0, 'limit': limit}),
            # v3 files API
            (f'https://drive.kdocs.cn/api/v3/files',
             {'groupid': self.teamid, 'parentid': parentid, 'start': 0, 'limit': limit}),
            # v2 API
            (f'https://drive.kdocs.cn/api/v2/groups/{self.teamid}/files',
             {'parentid': parentid, 'start': 0, 'limit': limit}),
        ]
        for url, params in api_paths:
            api_name = url.split('/api/')[1].split('?')[0] if '/api/' in url else url
            try:
                r = self.session.get(url, params=params, timeout=10, allow_redirects=False)
                # 重定向说明cookie失效了，直接跳过
                if r.status_code in (301, 302):
                    print(f"[kdocs] API {api_name} 重定向({r.status_code})，cookie可能已失效")
                    continue
                if r.status_code == 200:
                    try:
                        data = r.json()
                    except Exception:
                        print(f"[kdocs] API {api_name} 返回非JSON内容，cookie可能失效")
                        continue
                    files = data.get('files', []) or data.get('data', {}).get('files', []) or data.get('items', []) or []
                    if files:
                        print(f"[kdocs] API成功: {api_name} -> {len(files)}个文件")
                        return files
                    # 即使空也可能是正确路径
                    print(f"[kdocs] API {api_name} 返回0个文件 (code={r.status_code})")
                else:
                    print(f"[kdocs] API {api_name} 返回状态码 {r.status_code}")
            except Exception as e:
                print(f"[kdocs] API {api_name} 错误: {e}")
        return []

    def get_download_url(self, file_id):
        """获取文件下载链接"""
        if not self.teamid:
            return None
        try:
            url = f'https://drive.kdocs.cn/api/v3/groups/{self.teamid}/files/{file_id}/download'
            r = self.session.get(url, timeout=10, allow_redirects=False)
            if r.status_code in (301, 302):
                print(f"[kdocs] get_download_url 重定向({r.status_code})，cookie可能已失效")
                return None
            if r.status_code == 200:
                try:
                    data = r.json()
                    return data.get('fileinfo', {}).get('url')
                except Exception:
                    print(f"[kdocs] get_download_url 返回非JSON")
                    return None
        except Exception as e:
            print(f"[kdocs] get_download_url error: {e}")
        return None

    def download_xlsx(self, file_id, save_path=None):
        """下载xlsx文件到本地，返回路径（加重定向检测+xlsx校验）"""
        import tempfile
        try:
            url = self.get_download_url(file_id)
            if not url:
                print(f"[kdocs] download_xlsx: no download URL for file_id={file_id}")
                return None
            try:
                r = self.session.get(url, timeout=30, allow_redirects=False)
                # 重定向说明跳登录了，cookie失效
                if r.status_code in (301, 302):
                    print(f"[kdocs] download_xlsx: file_id={file_id} 重定向({r.status_code})，cookie可能已失效")
                    return None
                if r.status_code == 200:
                    # 验证是否是真实的xlsx文件（前2个字节是PK）
                    if len(r.content) < 100 or r.content[:2] != b'PK':
                        print(f"[kdocs] download_xlsx: file_id={file_id} 内容不是xlsx({len(r.content)}bytes)，cookie可能已失效")
                        return None
                    if save_path is None:
                        save_path = os.path.join(tempfile.gettempdir(), f'kdocs_{file_id}.xlsx')
                    with open(save_path, 'wb') as f:
                        f.write(r.content)
                    print(f"[kdocs] downloaded: {len(r.content)} bytes -> {os.path.basename(save_path)}")
                    return save_path
                else:
                    print(f"[kdocs] download_xlsx: HTTP {r.status_code} for file_id={file_id}")
            except Exception as e:
                print(f"[kdocs] download_xlsx error: {e}")
        except Exception as e:
            print(f"[kdocs] download_xlsx outer error: {e}")
        return None

    def read_sheet(self, file_id=None, xlsx_path=None, sheet_name=None):
        """读取xlsx文件内容，返回二维数组"""
        import openpyxl
        if not xlsx_path and file_id:
            xlsx_path = self.download_xlsx(file_id)
        if not xlsx_path:
            return []
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            wb.close()
            return rows
        except Exception as e:
            print(f"[kdocs] read_sheet error: {e}")
            return []

    # ============================================================
    # 旧版方法（分享链接方式，兼容保留）
    # ============================================================
    def get_folder_files(self):
        """获取分享文件夹中的所有文件列表（旧版兼容）"""
        return self.list_team_files()

    def get_file_content(self, file_id):
        """获取文件内容（旧版兼容，返回xlsx路径）"""
        return self.download_xlsx(file_id)
